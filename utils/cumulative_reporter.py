from __future__ import annotations

import importlib.util
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from utils.excel_reporter import (
    add_bug_report_sheet,
    add_summary_sheet,
    build_bug_report_row,
    build_sheet,
    load_artifacts_for_xml,
    resolve_meta_for_test,
)
from utils.testcase_identity import canonical_token


PROJECT_ROOT = Path(__file__).resolve().parents[1]
REPORTS_DIR = PROJECT_ROOT / "reports"
CATALOG_META_PATH = PROJECT_ROOT / "catalog" / "test_case_meta.py"


def _load_test_case_meta() -> dict:
    spec = importlib.util.spec_from_file_location("cumulative_test_case_meta", CATALOG_META_PATH)
    if spec is None or spec.loader is None:
        return {}
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return getattr(module, "TEST_CASE_META", {})


def _parse_status_and_error(testcase: ET.Element) -> tuple[str, str]:
    failure = testcase.find("failure")
    error = testcase.find("error")
    skipped = testcase.find("skipped")

    if failure is not None:
        return "failed", (failure.get("message", "") or (failure.text or "")).strip()
    if error is not None:
        return "failed", (error.get("message", "") or (error.text or "")).strip()
    if skipped is not None:
        return "skipped", ""
    return "passed", ""


def _format_run_date_from_filename(xml_path: Path) -> str:
    stem = xml_path.stem
    if stem.startswith("results_"):
        try:
            dt = datetime.strptime(stem.removeprefix("results_"), "%Y%m%d_%H%M%S")
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            pass
    return datetime.fromtimestamp(xml_path.stat().st_mtime).strftime("%d/%m/%Y")


def build_latest_results(reports_dir) -> dict:
    reports_dir = Path(reports_dir)
    latest_results: dict[str, dict] = {}

    for xml_path in sorted(reports_dir.glob("results_*.xml"), key=lambda item: item.name):
        try:
            root = ET.parse(xml_path).getroot()
        except ET.ParseError:
            continue

        run_date_display = _format_run_date_from_filename(xml_path)
        screenshots = load_artifacts_for_xml(xml_path)

        # Works for both <testsuites><testsuite>... and bare <testsuite> roots.
        for testcase in root.iter("testcase"):
            classname = testcase.get("classname", "")
            name = testcase.get("name", "")
            if not classname or not name:
                continue
            if "demo_failures" in classname:
                continue

            status, error_text = _parse_status_and_error(testcase)
            meta_key = f"{classname}::{name}"
            latest_results[meta_key] = {
                "status": status,
                "error_text": error_text,
                "run_date_display": run_date_display,
                "xml_file": xml_path.name,
                "timestamp_token": xml_path.stem.removeprefix("results_"),
                "screenshot_path": (screenshots.get(meta_key) or {}).get("path"),
            }

    return latest_results


def _pick_latest_entry(entries: list[dict]) -> dict:
    return sorted(entries, key=lambda item: item.get("timestamp_token", ""))[-1]


def _resolve_meta_result(meta_key: str, latest_results: dict) -> dict:
    exact_match = latest_results.get(meta_key)
    if exact_match:
        return exact_match

    canonical_meta_key = canonical_token(meta_key)
    for result_key, result_value in latest_results.items():
        if canonical_token(result_key) == canonical_meta_key:
            return result_value

    classname, test_name = meta_key.split("::", 1)
    matching_entries = []
    for result_key, result_value in latest_results.items():
        result_classname, result_name = result_key.split("::", 1)
        if result_classname != classname:
            continue
        if result_name.startswith(test_name + "[") or result_name.startswith(test_name + " "):
            matching_entries.append(result_value)
            continue
        if canonical_token(result_name) == canonical_token(test_name):
            matching_entries.append(result_value)

    if not matching_entries:
        return {
            "status": "skipped",
            "error_text": "",
            "run_date_display": "",
            "xml_file": None,
            "timestamp_token": "",
        }

    failed_entries = [entry for entry in matching_entries if entry["status"] == "failed"]
    if failed_entries:
        return _pick_latest_entry(failed_entries)

    if all(entry["status"] == "passed" for entry in matching_entries):
        return _pick_latest_entry(matching_entries)

    return _pick_latest_entry(
        [entry for entry in matching_entries if entry["status"] == "passed"] or matching_entries
    ) if any(entry["status"] == "passed" for entry in matching_entries) else {
        "status": "skipped",
        "error_text": "",
        "run_date_display": "",
        "xml_file": None,
        "timestamp_token": "",
    }


def _build_case_entries(test_case_meta: dict, latest_results: dict) -> list[dict]:
    candidate_keys = set(latest_results)
    for meta_key in test_case_meta:
        if meta_key in latest_results:
            candidate_keys.add(meta_key)
            continue

        meta_classname, meta_name = meta_key.split("::", 1)
        has_parametrized_history = any(
            result_key.startswith(f"{meta_classname}::{meta_name}[") or result_key.startswith(f"{meta_classname}::{meta_name} ")
            for result_key in latest_results
        )
        if not has_parametrized_history:
            candidate_keys.add(meta_key)

    candidate_keys = sorted(candidate_keys)
    generated_meta_ids: dict[str, int] = {}
    seen_ids: dict[str, int] = {}
    entries: list[dict] = []

    for case_key in candidate_keys:
        classname, name = case_key.split("::", 1)
        meta = resolve_meta_for_test(classname, name, test_case_meta, generated_meta_ids)
        resolved = _resolve_meta_result(case_key, latest_results)

        base_id = meta["id"]
        if base_id in seen_ids:
            seen_ids[base_id] += 1
            case_id = f"{base_id}_{seen_ids[base_id]}"
        else:
            seen_ids[base_id] = 0
            case_id = base_id

        entries.append(
            {
                "lookup_key": case_key,
                "case_id": case_id,
                "meta": {**meta, "id": case_id},
                "result": resolved,
            }
        )

    return entries


def _format_pct(numerator: int, denominator: int):
    if denominator <= 0:
        return 0
    raw = round((numerator / denominator) * 100, 1)
    return int(raw) if raw.is_integer() else raw


def get_coverage_summary(latest_results: dict, test_case_meta: dict) -> dict:
    case_entries = _build_case_entries(test_case_meta, latest_results)
    total_cases = len(case_entries)
    executed_cases = 0
    passed_cases = 0
    failed_cases = 0

    for case_entry in case_entries:
        resolved = case_entry["result"]
        status = resolved["status"]
        if status != "skipped":
            executed_cases += 1
        if status == "passed":
            passed_cases += 1
        elif status == "failed":
            failed_cases += 1

    not_executed_cases = total_cases - executed_cases
    return {
        "total_cases": total_cases,
        "executed_cases": executed_cases,
        "not_executed_cases": not_executed_cases,
        "passed_cases": passed_cases,
        "failed_cases": failed_cases,
        "coverage_pct": _format_pct(executed_cases, total_cases),
    }


def generate_cumulative_report(output_path=None, reports_dir=REPORTS_DIR):
    reports_dir = Path(reports_dir)
    test_case_meta = _load_test_case_meta()
    latest_results = build_latest_results(reports_dir)

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = reports_dir / f"report_excel_cumulative_{timestamp}.xlsx"
    output_path = Path(output_path)

    workbook = Workbook()
    workbook.remove(workbook.active)

    case_entries = _build_case_entries(test_case_meta, latest_results)
    sheet_count = 0
    summary_rows: list[dict] = []
    bug_rows: list[dict] = []

    for case_entry in case_entries:
        meta = case_entry["meta"]
        resolved = case_entry["result"]
        sheet_name = case_entry["case_id"]
        worksheet = workbook.create_sheet(title=sheet_name)
        build_sheet(
            worksheet,
            meta,
            resolved["status"],
            resolved["error_text"],
            resolved["run_date_display"],
            summary_sheet_name="Tong_Hop_Tich_Luy",
            screenshot_path=resolved.get("screenshot_path"),
        )
        bug_row = build_bug_report_row(
            meta=meta,
            sheet_name=sheet_name,
            result=resolved["status"],
            error_text=resolved["error_text"],
            run_date=resolved["run_date_display"],
            xml_file=resolved["xml_file"],
            screenshot_path=resolved.get("screenshot_path"),
            classname=case_entry["lookup_key"].split("::", 1)[0],
        )
        if bug_row:
            bug_rows.append(bug_row)
        summary_rows.append(
            {
                "sheet_name": sheet_name,
                "case_id": case_entry["case_id"],
                "description": meta["description"],
                "result": resolved["status"],
                "run_date": resolved["run_date_display"],
                "xml_file": resolved["xml_file"],
            }
        )
        sheet_count += 1

    if sheet_count == 0:
        worksheet = workbook.create_sheet(title="Khong_Co_Test_Map")
        worksheet["A1"] = "Khong tim thay test case nao trong catalog/test_case_meta.py."
    else:
        add_summary_sheet(workbook, summary_rows, title="Tong_Hop_Tich_Luy")
        add_bug_report_sheet(workbook, bug_rows, title="Bug_Report", summary_sheet_name="Tong_Hop_Tich_Luy")

    workbook.save(output_path)
    print(f"[Cumulative Excel] {sheet_count} sheet(s) -> {output_path}")
    return output_path


if __name__ == "__main__":
    if hasattr(__import__("sys").stdout, "reconfigure"):
        __import__("sys").stdout.reconfigure(encoding="utf-8")
    generated = generate_cumulative_report(reports_dir=REPORTS_DIR)
    meta = _load_test_case_meta()
    latest = build_latest_results(REPORTS_DIR)
    summary = get_coverage_summary(latest, meta)
    print(f"[Cumulative Excel] Coverage: {summary['executed_cases']}/{summary['total_cases']} ({summary['coverage_pct']}%)")
    print(f"[Cumulative Excel] Saved: {generated}")
