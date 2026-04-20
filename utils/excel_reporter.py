"""
Sinh bao cao Excel cho ket qua pytest.
Moi test case duoc xuat thanh mot sheet, theo bo cuc mau dang bang test case.

Usage:
    python -m utils.excel_reporter
    python -m utils.excel_reporter reports/results_YYYYMMDD_HHMMSS.xml
"""

from __future__ import annotations

import json
import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

from utils.testcase_identity import canonical_token, normalize_test_text


DARK_BLUE = "1F3864"
MED_BLUE = "2E74B5"
YELLOW = "FFFF00"
LIGHT_GREEN = "E2EFDA"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
FAIL_RED = "FFE0E0"
BLACK = "000000"

TESTER_NAME = "Selenium Auto"

RESULT_LABELS = {
    "passed": "Pass",
    "failed": "Fail",
    "skipped": "Not executed",
}

SUMMARY_SHEET_NAMES = {"Tong_Hop_Test_Case", "Tong_Hop_Tich_Luy", "Bug_Report"}

MODULE_PREFIX_MAP = {
    "tests.test_login": "DN",
    "tests.test_register_module": "DK",
    "tests.test_account_module": "GDTK",
    "tests.test_cart_checkout_module": "GH",
    "tests.test_runtime_audit": "AUD",
    "tests.test_source_audit": "SRC",
    "tests.test_source_conformance": "SCF",
    "tests.test_search": "TK",
    "tests.test_pages": "TRANG",
    "tests.test_home_filters": "LOC",
    "tests.test_home_products": "SP",
    "tests.test_ui": "UI",
    "tests.test_catalog_integrity": "META",
}

TAB_COLORS = {
    "passed": "70AD47",
    "failed": "C00000",
    "skipped": "FFC000",
}

TECHNIQUE_TRANSLATIONS = {
    "Functional Testing": "Kiểm thử chức năng",
    "Usability Testing": "Kiểm thử khả dụng",
    "Navigation Testing": "Kiểm thử điều hướng",
    "UI / Layout Testing": "Kiểm thử giao diện / bố cục",
    "UI / Asset Audit": "Kiểm thử giao diện / tài nguyên",
    "Functional / Validation": "Kiểm thử chức năng / kiểm tra hợp lệ",
    "Business Flow / Regression": "Kiểm thử luồng nghiệp vụ / hồi quy",
    "Meta Validation": "Kiểm tra tính hợp lệ metadata",
    "Negative / Demo": "Kiểm thử âm / demo",
    "Regression / End-to-End": "Kiểm thử hồi quy / end-to-end",
    "Black-box / Validation / UI": "Kiểm thử hộp đen / kiểm tra hợp lệ / giao diện",
    "Black-box / Equivalence / Boundary": "Kiểm thử hộp đen / phân lớp tương đương / giá trị biên",
    "Black-box / Validation / Session": "Kiểm thử hộp đen / kiểm tra hợp lệ / phiên đăng nhập",
    "Navigation / Access Check": "Kiểm tra điều hướng / quyền truy cập",
    "UI / Usability / Smoke": "Kiểm thử giao diện / khả dụng / smoke",
    "Quality Gate / Deployment Audit": "Kiểm thử chất lượng / audit triển khai",
    "Runtime Audit / Browser Console": "Kiểm thử runtime / nhật ký trình duyệt",
    "Runtime Audit / Business Flow": "Kiểm thử runtime / luồng nghiệp vụ",
    "Deployment / Deep-Link Audit": "Kiểm thử triển khai / liên kết sâu",
    "Equivalence Partitioning (Valid)": "Phân lớp tương đương (hợp lệ)",
    "Equivalence Partitioning (Invalid)": "Phân lớp tương đương (không hợp lệ)",
    "Boundary Value Analysis (Valid)": "Phân tích giá trị biên (hợp lệ)",
    "Boundary Value Analysis (Invalid)": "Phân tích giá trị biên (không hợp lệ)",
}

TEXT_TRANSLATIONS = {
    "Homepage should not emit SEVERE browser console errors during load": "Trang chủ không được phát sinh lỗi SEVERE trên console khi tải trang",
    "Login page should not emit SEVERE browser console errors": "Trang đăng nhập không được phát sinh lỗi SEVERE trên console",
    "Visible homepage images should load without broken asset state": "Các ảnh hiển thị trên trang chủ phải tải được, không bị lỗi tài nguyên",
    "Homepage same-origin links should be directly reachable without 404": "Các liên kết cùng domain trên trang chủ phải truy cập trực tiếp được, không trả về 404",
    "Checkout page with seeded session should not emit SEVERE browser console errors": "Trang thanh toán với phiên đã seed không được phát sinh lỗi SEVERE trên console",
    "Website is reachable": "Truy cập được website",
    "Authenticated test account and seeded checkout data are available": "Có tài khoản test đã đăng nhập và dữ liệu checkout đã được seed",
    "Authenticated user via API": "Người dùng đã xác thực qua API",
    "Checkout data seeded in localStorage": "Dữ liệu checkout đã được seed vào localStorage",
    "Open homepage": "Mở trang chủ",
    "Homepage loads successfully": "Trang chủ tải thành công",
    "Read browser console logs": "Đọc nhật ký console của trình duyệt",
    "No SEVERE entries are present": "Không có bản ghi SEVERE",
    "Open login page through UI navigation": "Mở trang đăng nhập thông qua điều hướng giao diện",
    "Login form is visible": "Form đăng nhập hiển thị",
    "Inspect visible images": "Kiểm tra các ảnh đang hiển thị",
    "No visible image has naturalWidth equal to zero": "Không có ảnh hiển thị nào có naturalWidth bằng 0",
    "Request each same-origin href": "Gửi request tới từng liên kết cùng domain",
    "Each route responds with status code below 400": "Mỗi route đều phản hồi với mã trạng thái nhỏ hơn 400",
    "Load authenticated checkout state": "Thiết lập trạng thái checkout đã đăng nhập",
    "User session is active": "Phiên người dùng đang hoạt động",
    "Open SPA checkout route": "Mở route checkout của SPA",
    "Checkout page is visible": "Trang thanh toán hiển thị",
}


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(bold: bool = False, color: str = BLACK) -> Font:
    return Font(bold=bold, color=color, name="Arial", size=10)


def _align(wrap: bool = True) -> Alignment:
    return Alignment(wrap_text=wrap, vertical="center", horizontal="left")


def _style(cell, bg: str = WHITE, bold: bool = False, fg: str = BLACK) -> None:
    cell.fill = _fill(bg)
    cell.font = _font(bold=bold, color=fg)
    cell.alignment = _align()


def _mc(ws, cell_range: str, bg: str = WHITE, bold: bool = False, fg: str = BLACK, value=None) -> None:
    parts = cell_range.split(":")
    ws.merge_cells(cell_range)
    tl = ws[parts[0]]
    _style(tl, bg=bg, bold=bold, fg=fg)
    if value is not None:
        tl.value = value


def _sc(ws, addr: str, bg: str = WHITE, bold: bool = False, fg: str = BLACK, value=None) -> None:
    cell = ws[addr]
    _style(cell, bg=bg, bold=bold, fg=fg)
    if value is not None:
        cell.value = value


def _repair_mojibake(text: str) -> str:
    return normalize_test_text(text)


def _default_description(test_name: str) -> str:
    return test_name.replace("_", " ").strip().title()


def _normalize_text(value) -> str:
    if value is None:
        return ""
    text = normalize_test_text(str(value))
    return TEXT_TRANSLATIONS.get(text, text)


def _truncate_text(text: str, max_len: int = 500) -> str:
    text = " ".join(_normalize_text(text).split())
    if len(text) <= max_len:
        return text
    return text[: max_len - 3].rstrip() + "..."


def _strip_assert_prefix(text: str) -> str:
    cleaned = _normalize_text(text).strip()
    prefixes = (
        "E       AssertionError:",
        "AssertionError:",
        "E   AssertionError:",
    )
    for prefix in prefixes:
        if cleaned.startswith(prefix):
            return cleaned.removeprefix(prefix).strip()
    return cleaned


def _extract_actionable_error(error_text: str) -> str:
    normalized = _normalize_text(error_text).strip()
    if not normalized:
        return "Không có thông báo lỗi chi tiết từ pytest."

    useful_lines = []
    for raw_line in normalized.splitlines():
        line = _strip_assert_prefix(raw_line)
        if not line:
            continue
        if line.startswith(("assert ", "where ", "+", "-", ">")):
            continue
        if "AssertionError:" in line:
            line = line.split("AssertionError:", 1)[1].strip()
        useful_lines.append(line)

    return _truncate_text(useful_lines[0] if useful_lines else normalized, 700)


def _translate_known_error(raw: str) -> str:
    normalized = _normalize_text(raw)
    known_messages = {
        "REG_13 expects register input to be trimmed before submit/validation.": (
            "REG_13 không đạt: dữ liệu nhập ở form đăng ký phải được trim khoảng trắng đầu/cuối "
            "trước khi submit hoặc validation, nhưng hệ thống/source hiện tại chưa thể hiện xử lý này."
        ),
        "CHK_16 expects checkout text fields to be trimmed.": (
            "CHK_16 không đạt: họ tên/SĐT/địa chỉ ở checkout phải được trim khoảng trắng đầu/cuối "
            "trước khi tạo đơn, nhưng hệ thống/source hiện tại chưa thể hiện xử lý này."
        ),
        "CRT_14 expects discounted price logic in cart/product flow.": (
            "CRT_14 không đạt: luồng product/cart cần có logic tính giá sau giảm giá, "
            "nhưng chưa tìm thấy dấu hiệu xử lý discount/sale/finalPrice."
        ),
        "VCH_06 expects voucher to allow total equal to minOrderValue.": (
            "VCH_06 không đạt: voucher phải được chấp nhận khi tổng tiền bằng đúng giá trị tối thiểu, "
            "nhưng source chưa thể hiện điều kiện >= hoặc tương đương."
        ),
        "VCH_07 expects voucher usage-limit validation.": (
            "VCH_07 không đạt: voucher cần kiểm tra giới hạn lượt sử dụng, "
            "nhưng source chưa thể hiện usageLimit/usedCount/limit."
        ),
        "VCH_08 expects voucher usage count to prevent repeated overuse.": (
            "VCH_08 không đạt: hệ thống cần cập nhật và kiểm soát số lần dùng voucher để tránh vượt lượt, "
            "nhưng source chưa thể hiện cơ chế tăng usedCount hoặc kiểm tra usageLimit."
        ),
    }
    for english, vietnamese in known_messages.items():
        if english in normalized:
            return vietnamese
    return normalized


def _vietnamese_error_summary(error_text: str, description: str, failed_step: dict | None) -> dict:
    raw = _translate_known_error(_extract_actionable_error(error_text))
    expected = _normalize_text((failed_step or {}).get("expected", ""))
    action = _normalize_text((failed_step or {}).get("detail", ""))

    lower_raw = raw.lower()
    if "expects" in lower_raw:
        actual = (
            "Điều kiện kiểm tra chưa đạt. Hệ thống/source hiện tại chưa chứng minh được "
            f"yêu cầu này: {raw}"
        )
    elif "expected" in lower_raw and "got" in lower_raw:
        actual = f"Giá trị thực tế khác giá trị mong đợi: {raw}"
    elif "got status=" in lower_raw or "status=" in lower_raw:
        actual = f"API trả về trạng thái không đúng yêu cầu testcase: {raw}"
    elif "timeout" in lower_raw or "timed out" in lower_raw:
        actual = f"Hệ thống phản hồi quá lâu hoặc request bị timeout: {raw}"
    elif "not found" in lower_raw or "missing" in lower_raw:
        actual = f"Thiếu thành phần/route/dữ liệu cần có theo testcase: {raw}"
    else:
        actual = raw

    expected_sentence = (expected or "xem cột Kết quả mong đợi").rstrip(".")
    summary = (
        f"Testcase '{description}' không đạt tại bước đối chiếu cuối. "
        f"Yêu cầu mong đợi: {expected_sentence}. "
        f"Kết quả thực tế: {actual}"
    )
    retest = (
        "Test lại thủ công theo đúng dữ liệu đầu vào, quan sát thông báo trên giao diện/API, "
        "sau đó đối chiếu với kết quả mong đợi. Nếu vẫn ra kết quả như cột 'Dấu hiệu lỗi hiện tại' "
        "thì ghi nhận đây là bug cần sửa."
    )

    return {
        "summary": _truncate_text(summary, 900),
        "actual": _truncate_text(actual, 500),
        "expected": expected,
        "action": action,
        "raw": _truncate_text(raw, 900),
        "retest": retest,
    }


def _normalize_list(items: list[str] | None) -> list[str]:
    if not items:
        return []
    return [_normalize_text(item) for item in items]


def _normalize_steps(steps: list[dict] | None) -> list[dict]:
    normalized_steps = []
    for step in steps or []:
        normalized_steps.append(
            {
                "detail": _normalize_text(step.get("detail", "")),
                "expected": _normalize_text(step.get("expected", "")),
            }
        )
    return normalized_steps


def _case_prefix(case_id: str) -> str:
    return case_id.split("_", 1)[0].upper() if case_id else ""


def _project_case_context(case_id: str) -> tuple[str, str, str]:
    prefix = _case_prefix(case_id)
    contexts = {
        "REG": (
            "Tại trang Đăng ký",
            'Click vào nút "Đăng ký" (Submit).',
            "Form Đăng ký hiển thị đầy đủ, các trường cần kiểm thử nhập được dữ liệu.",
        ),
        "LOG": (
            "Tại trang Đăng nhập",
            'Click vào nút "Đăng nhập" (Submit).',
            "Form Đăng nhập hiển thị đầy đủ, ô tài khoản và mật khẩu nhập được dữ liệu.",
        ),
        "PWD": (
            "Tại trang Thay đổi mật khẩu",
            'Click vào nút "Lưu" hoặc "Đổi mật khẩu" (Submit).',
            "Form Thay đổi mật khẩu hiển thị đầy đủ, các ô mật khẩu nhập được dữ liệu.",
        ),
        "CRT": (
            "Tại trang Chi tiết sản phẩm hoặc trang Giỏ hàng",
            "Thực hiện thao tác giỏ hàng theo dữ liệu kiểm thử.",
            "Trang sản phẩm/giỏ hàng hiển thị đầy đủ sản phẩm, thuộc tính, số lượng và tổng tiền cần kiểm thử.",
        ),
        "CHK": (
            "Tại trang Checkout / Thanh toán",
            'Click vào nút "Xác nhận đặt hàng" hoặc thao tác thanh toán theo testcase.',
            "Trang Checkout hiển thị đầy đủ thông tin nhận hàng, danh sách sản phẩm, voucher và phương thức thanh toán.",
        ),
        "VCH": (
            "Tại khu vực Voucher của trang Checkout / Thanh toán",
            'Click vào nút "Áp dụng" voucher.',
            "Ô nhập voucher và thông tin tổng tiền hiển thị đầy đủ để kiểm thử điều kiện áp dụng mã giảm giá.",
        ),
    }
    return contexts.get(prefix, ("", "", ""))


def _specific_expected(steps: list[dict]) -> str:
    generic_markers = (
        "kết quả thực tế",
        "đúng module",
        "đúng ngữ cảnh",
        "report phải",
        "report có",
        "ghi rõ bug",
        "bằng chứng lỗi",
        "bug finding liên quan",
    )
    for step in reversed(steps):
        expected = _normalize_text(step.get("expected", "")).strip()
        if not expected:
            continue
        lower_expected = expected.lower()
        if any(marker in lower_expected for marker in generic_markers):
            continue
        return expected
    return steps[-1].get("expected", "") if steps else ""


def _project_full_data_lines(case_id: str, action: str = "") -> list[str]:
    prefix = _case_prefix(case_id)
    action = _normalize_text(action).strip().rstrip(".")

    if prefix == "REG":
        username = "tien_user_001"
        phone = "0912345678"
        password = "Abcd123@"
        confirm_password = password
        if case_id == "REG_13":
            username = " user123 "
            phone = " 0912345678 "
        elif case_id == "REG_14":
            phone = "0987 654 321"
        elif case_id == "REG_15":
            password = "Abc@ 1234"
            confirm_password = "Abc@ 1234"
        elif case_id == "REG_16":
            username = "user@@@"
        elif case_id == "REG_17":
            username = "u" * 260
        return [
            "Họ tên: Nguyễn Văn A",
            f"Tên đăng nhập: {username}",
            "Email nếu form có: tien.user001@mail.test",
            f"Số điện thoại: {phone}",
            f"Mật khẩu: {password}",
            f"Nhập lại mật khẩu: {confirm_password}",
            f"Điểm đang kiểm tra: {action or 'dữ liệu đăng ký hợp lệ'}",
        ]

    if prefix == "LOG":
        account = "tien_user_001"
        password = "Abcd123@"
        if case_id == "LOG_10":
            account = " user123 "
        elif case_id == "LOG_11":
            password = " pass123 "
        return [
            f"Tài khoản / Email / SĐT: {account}",
            f"Mật khẩu: {password}",
            "Remember me: không tick, trừ testcase yêu cầu kiểm tra remember me",
            f"Điểm đang kiểm tra: {action or 'đăng nhập theo dữ liệu hợp lệ'}",
        ]

    if prefix == "PWD":
        current_password = "Abcd123@"
        new_password = "Xyzc123@"
        token_state = "Có token đăng nhập hợp lệ"
        if case_id == "PWD_02":
            current_password = ""
            new_password = ""
        if case_id == "PWD_03":
            current_password = "SaiPass@123"
        if case_id == "PWD_06":
            token_state = "Không có token đăng nhập"
        if case_id == "PWD_05":
            new_password = "abc"
        if case_id in {"PWD_04", "PWD_07"}:
            new_password = "Abcd123@"
        return [
            "User test: tien_user_001",
            f"Mật khẩu hiện tại: {current_password or '(bỏ trống)'}",
            f"Mật khẩu mới: {new_password}",
            f"Nhập lại mật khẩu mới: {new_password}",
            f"Trạng thái đăng nhập/token: {token_state}",
            f"Điểm đang kiểm tra: {action or 'luồng đổi mật khẩu'}",
        ]

    if prefix == "CRT":
        quantity = "1"
        if case_id == "CRT_12":
            quantity = "999"
        return [
            "Sản phẩm: Giày thể thao test / sản phẩm nổi bật đầu tiên",
            "Size: 40",
            "Màu sắc: Đen",
            f"Số lượng: {quantity}",
            "Giá tham chiếu: lấy từ sản phẩm thực tế trên web/API",
            f"Điểm đang kiểm tra: {action or 'luồng thêm/cập nhật giỏ hàng'}",
        ]

    if prefix in {"CHK", "VCH"}:
        voucher = ""
        payment_method = "COD"
        phone = "0912345678"
        items = "1 sản phẩm, size 40, màu Đen, số lượng 1"
        if case_id == "CHK_17":
            phone = "09abc12345"
        if case_id == "CHK_18":
            payment_method = "null"
        if case_id == "CHK_20":
            items = "[]"
        if case_id in {"CHK_06", "VCH_06"}:
            voucher = "GIAM100K"
        elif case_id in {"CHK_07", "VCH_07", "VCH_08"}:
            voucher = "SAIBET"
        elif case_id == "CHK_08":
            voucher = "EXPIRED100K"
        elif case_id == "CHK_09":
            voucher = "MIN500K"
        return [
            "Họ tên người nhận: Nguyễn Văn A",
            f"Số điện thoại: {phone}",
            "Địa chỉ: 123 Lê Lợi, Phường Bến Nghé, Quận 1, TP.HCM",
            f"Sản phẩm trong giỏ: {items}",
            f"Phương thức thanh toán: {payment_method}",
            f"Voucher: {voucher or 'không nhập'}",
            f"Điểm đang kiểm tra: {action or 'luồng checkout/voucher'}",
        ]

    return [action] if action else []


def _project_full_data_block(case_id: str, action: str = "") -> str:
    lines = _project_full_data_lines(case_id, action)
    return "\n".join(f"- {line}" for line in lines if line)


def _project_test_data_rows(meta: dict, original_rows: list[str]) -> list[str]:
    case_id = _normalize_text(meta.get("id", "")).strip()
    if not _project_case_context(case_id)[0]:
        return original_rows
    action = "; ".join(original_rows)
    return _project_full_data_lines(case_id, action)


def _project_manual_steps(meta: dict, steps: list[dict]) -> list[dict]:
    case_id = _normalize_text(meta.get("id", "")).strip()
    context, submit_action, first_expected = _project_case_context(case_id)
    if not context:
        return steps

    description = _normalize_text(meta.get("description", "")).strip().rstrip(".")
    test_data = [_normalize_text(item).strip() for item in meta.get("test_data", []) if _normalize_text(item).strip()]
    action = "; ".join(test_data).strip().rstrip(".")
    if not action and len(steps) > 1:
        action = _normalize_text(steps[1].get("detail", "")).strip().rstrip(".")
    if not action:
        action = "Thực hiện đúng dữ liệu đầu vào theo testcase"

    data_block = _project_full_data_block(case_id, action)
    if not data_block:
        data_block = f"- {action}"

    expected = _specific_expected(steps).strip().rstrip(".")
    if not expected:
        expected = "Kết quả hiển thị đúng theo đặc tả testcase"

    return [
        {
            "detail": (
                f"{context}, nhập hoặc thiết lập dữ liệu kiểm thử vào form/chức năng. "
                "Dữ liệu nhập đầy đủ:\n"
                f"{data_block}"
            ),
            "expected": first_expected,
        },
        {
            "detail": submit_action,
            "expected": (
                f"- Hệ thống xử lý đúng theo testcase {case_id}: {description}.\n"
                f"- {expected}.\n"
                "- Không lưu, không cập nhật hoặc không chuyển trạng thái sai ngoài yêu cầu testcase."
            ),
        },
    ]


def _translate_technique(value: str | None) -> str:
    normalized = _normalize_text(value or "Kiểm thử chức năng")
    return TECHNIQUE_TRANSLATIONS.get(normalized, normalized)


def _add_back_to_summary_link(ws, summary_sheet_name: str) -> None:
    ws.column_dimensions["Q"].width = 18
    ws.column_dimensions["R"].width = 18
    _mc(
        ws,
        "Q1:R2",
        bg=MED_BLUE,
        bold=True,
        fg=WHITE,
        value="Quay lai\nTong hop",
    )
    link_cell = ws["Q1"]
    link_cell.hyperlink = f"#'{summary_sheet_name}'!A1"
    link_cell.style = "Hyperlink"
    link_cell.alignment = Alignment(
        wrap_text=True,
        vertical="center",
        horizontal="center",
    )
    link_cell.font = Font(
        bold=True,
        color="FFFFFF",
        name="Arial",
        size=11,
        underline="single",
    )


def build_sheet(
    ws,
    meta: dict,
    result: str,
    error_text: str,
    run_date: str,
    summary_sheet_name: str = "Tong_Hop_Test_Case",
    screenshot_path: str | None = None,
) -> None:
    tc_id = meta["id"]
    description = _normalize_text(meta["description"])
    technique = _translate_technique(meta.get("technique"))
    prereqs = _normalize_list(meta.get("prerequisites", []))
    test_data = _normalize_list(meta.get("test_data", []))
    test_data = _project_test_data_rows(meta, test_data)
    steps = _normalize_steps(meta.get("steps", []))
    steps = _project_manual_steps(meta, steps)
    failed_step = steps[-1] if steps else {}
    failure_detail = _vietnamese_error_summary(error_text, description, failed_step) if result == "failed" else {}

    if result == "passed":
        overall_label = RESULT_LABELS["passed"]
        overall_bg = LIGHT_GREEN
    elif result == "failed":
        overall_label = RESULT_LABELS["failed"]
        overall_bg = FAIL_RED
    else:
        overall_label = RESULT_LABELS["skipped"]
        overall_bg = WHITE

    ws.sheet_properties.tabColor = TAB_COLORS.get(result, TAB_COLORS["skipped"])

    widths = {
        "A": 7,
        "B": 24,
        "C": 10,
        "D": 10,
        "E": 10,
        "F": 7,
        "G": 24,
        "H": 10,
        "I": 10,
        "J": 24,
        "K": 12,
        "L": 10,
        "M": 12,
        "N": 10,
        "O": 10,
        "P": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24
    _add_back_to_summary_link(ws, summary_sheet_name)

    _mc(
        ws,
        "A1:P1",
        bg=DARK_BLUE,
        bold=True,
        fg=WHITE,
        value=f"Mã ca kiểm thử      {tc_id}      Mô tả ca kiểm thử      {description}",
    )

    _mc(ws, "A2:B2", bg=MED_BLUE, bold=True, fg=WHITE, value="Người tạo")
    _mc(ws, "C2:E2", bg=WHITE, value=TESTER_NAME)
    _mc(ws, "F2:G2", bg=MED_BLUE, bold=True, fg=WHITE, value="Người rà soát")
    _mc(ws, "H2:J2", bg=WHITE, value="")
    _mc(ws, "K2:L2", bg=MED_BLUE, bold=True, fg=WHITE, value="Phiên bản")
    _mc(ws, "M2:P2", bg=WHITE, value="1.0")

    _mc(ws, "A3:C3", bg=WHITE, value="")
    _mc(ws, "D3:E3", bg=MED_BLUE, bold=True, fg=WHITE, value="Kỹ thuật")
    _mc(ws, "F3:P3", bg=YELLOW, value=technique)

    _mc(ws, "A4:P4", bg=DARK_BLUE, bold=True, fg=WHITE, value="Nhật ký kiểm thử QA")
    _mc(ws, "A5:P5", bg=DARK_BLUE, value="")

    _mc(ws, "A6:B6", bg=MED_BLUE, bold=True, fg=WHITE, value="Người kiểm thử")
    _mc(ws, "C6:E6", bg=WHITE, value=TESTER_NAME)
    _mc(ws, "F6:G6", bg=MED_BLUE, bold=True, fg=WHITE, value="Ngày kiểm thử")
    _mc(ws, "H6:J6", bg=WHITE, value=run_date)
    _mc(ws, "K6:P6", bg=overall_bg, bold=True, value=f"Kết quả ca kiểm thử: {overall_label}")

    _mc(ws, "A7:P7", bg=WHITE, value="")

    _sc(ws, "A8", bg=MED_BLUE, bold=True, fg=WHITE, value="STT")
    _mc(ws, "B8:E8", bg=MED_BLUE, bold=True, fg=WHITE, value="Điều kiện tiên quyết")
    _sc(ws, "F8", bg=MED_BLUE, bold=True, fg=WHITE, value="STT")
    _mc(ws, "G8:P8", bg=MED_BLUE, bold=True, fg=WHITE, value="Dữ liệu kiểm thử")

    max_pt = max(len(prereqs), len(test_data), 1)
    for index in range(max_pt):
        row = 9 + index
        bg = LIGHT_GRAY if index % 2 == 0 else WHITE
        ws.row_dimensions[row].height = 18
        _sc(ws, f"A{row}", bg=bg, value=(index + 1) if index < len(prereqs) else "")
        _mc(ws, f"B{row}:E{row}", bg=bg, value=prereqs[index] if index < len(prereqs) else "")
        _sc(ws, f"F{row}", bg=bg, value=(index + 1) if index < len(test_data) else "")
        _mc(ws, f"G{row}:P{row}", bg=bg, value=test_data[index] if index < len(test_data) else "")

    next_row = 9 + max_pt
    _mc(ws, f"A{next_row}:P{next_row}", bg=WHITE, value="")

    scenario_row = next_row + 1
    _mc(ws, f"A{scenario_row}:P{scenario_row}", bg=DARK_BLUE, bold=True, fg=WHITE, value="Kịch bản kiểm thử")

    empty_row = scenario_row + 1
    _mc(ws, f"A{empty_row}:P{empty_row}", bg=WHITE, value="")

    header_row = empty_row + 1
    ws.row_dimensions[header_row].height = 30
    _sc(ws, f"A{header_row}", bg=MED_BLUE, bold=True, fg=WHITE, value="Bước #")
    _mc(ws, f"B{header_row}:F{header_row}", bg=MED_BLUE, bold=True, fg=WHITE, value="Chi tiết bước")
    _mc(ws, f"G{header_row}:I{header_row}", bg=MED_BLUE, bold=True, fg=WHITE, value="Kết quả mong đợi")
    _mc(ws, f"J{header_row}:L{header_row}", bg=MED_BLUE, bold=True, fg=WHITE, value="Kết quả thực tế")
    _mc(
        ws,
        f"M{header_row}:P{header_row}",
        bg=MED_BLUE,
        bold=True,
        fg=WHITE,
        value="Pass / Fail / Not executed / Suspended / Crashed",
    )

    for index, step in enumerate(steps):
        row = header_row + 1 + index
        bg = LIGHT_GRAY if index % 2 == 0 else WHITE
        step_text_size = len(step.get("detail", "")) + len(step.get("expected", ""))
        has_multiline_step = "\n" in step.get("detail", "") or "\n" in step.get("expected", "")
        ws.row_dimensions[row].height = 90 if has_multiline_step else (58 if step_text_size > 180 else 34)

        if result == "passed":
            step_result = RESULT_LABELS["passed"]
            result_bg = LIGHT_GREEN
            actual = "Như mong đợi"
        elif result == "failed":
            if index == len(steps) - 1:
                step_result = RESULT_LABELS["failed"]
                result_bg = FAIL_RED
                actual = failure_detail.get("actual") or "Không như mong đợi"
            else:
                step_result = RESULT_LABELS["passed"]
                result_bg = LIGHT_GREEN
                actual = "Như mong đợi"
        else:
            step_result = RESULT_LABELS["skipped"]
            result_bg = WHITE
            actual = ""

        _sc(ws, f"A{row}", bg=bg, value=index + 1)
        _mc(ws, f"B{row}:F{row}", bg=bg, value=step.get("detail", ""))
        _mc(ws, f"G{row}:I{row}", bg=bg, value=step.get("expected", ""))
        _mc(ws, f"J{row}:L{row}", bg=bg, value=actual)
        _mc(ws, f"M{row}:P{row}", bg=result_bg, bold=True, value=step_result)

    current_row = header_row + len(steps) + 1

    if result == "failed":
        summary_title_row = current_row + 1
        summary_header_row = summary_title_row + 1
        summary_value_row = summary_header_row + 1
        _mc(
            ws,
            f"A{summary_title_row}:P{summary_title_row}",
            bg=DARK_BLUE,
            bold=True,
            fg=WHITE,
            value="Tóm tắt lỗi cần xử lý",
        )
        _mc(ws, f"A{summary_header_row}:D{summary_header_row}", bg=MED_BLUE, bold=True, fg=WHITE, value="Lỗi đang xảy ra")
        _mc(ws, f"E{summary_header_row}:H{summary_header_row}", bg=MED_BLUE, bold=True, fg=WHITE, value="Yêu cầu mong đợi")
        _mc(ws, f"I{summary_header_row}:L{summary_header_row}", bg=MED_BLUE, bold=True, fg=WHITE, value="Cách test lại")
        _mc(ws, f"M{summary_header_row}:P{summary_header_row}", bg=MED_BLUE, bold=True, fg=WHITE, value="Thông báo kỹ thuật gốc")
        ws.row_dimensions[summary_value_row].height = 72
        _mc(ws, f"A{summary_value_row}:D{summary_value_row}", bg=FAIL_RED, value=failure_detail.get("summary", "Testcase không đạt."))
        _mc(ws, f"E{summary_value_row}:H{summary_value_row}", bg=WHITE, value=failure_detail.get("expected", "Xem cột Kết quả mong đợi ở trên."))
        _mc(ws, f"I{summary_value_row}:L{summary_value_row}", bg=WHITE, value=failure_detail.get("retest", "Test lại theo đúng các bước ở bảng bên dưới."))
        _mc(ws, f"M{summary_value_row}:P{summary_value_row}", bg=FAIL_RED, value=failure_detail.get("raw", _short_error(error_text)))

        retest_title_row = summary_value_row + 2
        retest_header_row = retest_title_row + 1
        _mc(
            ws,
            f"A{retest_title_row}:P{retest_title_row}",
            bg=DARK_BLUE,
            bold=True,
            fg=WHITE,
            value="Hướng dẫn test lại thủ công để tái hiện lỗi",
        )
        _sc(ws, f"A{retest_header_row}", bg=MED_BLUE, bold=True, fg=WHITE, value="Bước #")
        _mc(ws, f"B{retest_header_row}:F{retest_header_row}", bg=MED_BLUE, bold=True, fg=WHITE, value="Thao tác test lại")
        _mc(ws, f"G{retest_header_row}:I{retest_header_row}", bg=MED_BLUE, bold=True, fg=WHITE, value="Kết quả mong đợi")
        _mc(ws, f"J{retest_header_row}:L{retest_header_row}", bg=MED_BLUE, bold=True, fg=WHITE, value="Dấu hiệu lỗi hiện tại")
        _mc(ws, f"M{retest_header_row}:P{retest_header_row}", bg=MED_BLUE, bold=True, fg=WHITE, value="Ghi chú")

        for index, step in enumerate(steps, start=1):
            row = retest_header_row + index
            bg = LIGHT_GRAY if index % 2 == 1 else WHITE
            step_text_size = len(step.get("detail", "")) + len(step.get("expected", ""))
            has_multiline_step = "\n" in step.get("detail", "") or "\n" in step.get("expected", "")
            ws.row_dimensions[row].height = 90 if has_multiline_step else (58 if step_text_size > 180 else 34)
            is_failure_point = index == len(steps)
            _sc(ws, f"A{row}", bg=bg, value=index)
            _mc(ws, f"B{row}:F{row}", bg=bg, value=step.get("detail", ""))
            _mc(ws, f"G{row}:I{row}", bg=bg, value=step.get("expected", ""))
            _mc(
                ws,
                f"J{row}:L{row}",
                bg=FAIL_RED if is_failure_point else bg,
                value=failure_detail.get("actual", _short_error(error_text)) if is_failure_point else "Như mong đợi",
            )
            _mc(
                ws,
                f"M{row}:P{row}",
                bg=FAIL_RED if is_failure_point else bg,
                bold=is_failure_point,
                value=(
                    "Bước này đang không đạt. Cần test lại và chụp/ghi nhận bằng chứng."
                    if is_failure_point
                    else "Tiếp tục theo đúng dữ liệu ở trên."
                ),
            )

        current_row = retest_header_row + len(steps)

    if screenshot_path and result == "failed":
        screenshot_title_row = current_row + 2
        screenshot_image_row = screenshot_title_row + 1
        _mc(
            ws,
            f"A{screenshot_title_row}:P{screenshot_title_row}",
            bg=DARK_BLUE,
            bold=True,
            fg=WHITE,
            value="Hinh anh loi",
        )
        ws.row_dimensions[screenshot_image_row].height = 260
        try:
            image = XLImage(screenshot_path)
            image.width = 780
            image.height = 420
            ws.add_image(image, f"A{screenshot_image_row}")
        except Exception:
            _mc(
                ws,
                f"A{screenshot_image_row}:P{screenshot_image_row}",
                bg=FAIL_RED,
                bold=True,
                value=f"Khong nhung duoc screenshot: {screenshot_path}",
            )


def add_summary_sheet(workbook, summary_rows: list[dict], title: str = "Tong_Hop_Test_Case") -> None:
    worksheet = workbook.create_sheet(title=title, index=0)
    worksheet.freeze_panes = "A5"
    worksheet.sheet_properties.tabColor = "5B9BD5"

    column_widths = {
        "A": 8,
        "B": 18,
        "C": 52,
        "D": 18,
        "E": 18,
        "F": 28,
        "G": 16,
    }
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    _mc(
        worksheet,
        "A1:G1",
        bg=DARK_BLUE,
        bold=True,
        fg=WHITE,
        value="Tổng hợp test case - Nhấn vào mã test case hoặc cột Mở chi tiết để nhảy đến đúng sheet báo cáo",
    )

    total_cases = len(summary_rows)
    passed_cases = sum(1 for row in summary_rows if row.get("result") == "passed")
    failed_cases = sum(1 for row in summary_rows if row.get("result") == "failed")
    skipped_cases = sum(1 for row in summary_rows if row.get("result") == "skipped")

    _mc(worksheet, "A2:B2", bg=MED_BLUE, bold=True, fg=WHITE, value="Tổng test case")
    _mc(worksheet, "C2:D2", bg=WHITE, bold=True, value=total_cases)
    _mc(worksheet, "E2:F2", bg=MED_BLUE, bold=True, fg=WHITE, value="Passed")
    _sc(worksheet, "G2", bg=LIGHT_GREEN, bold=True, value=passed_cases)

    _mc(worksheet, "A3:B3", bg=MED_BLUE, bold=True, fg=WHITE, value="Failed")
    _mc(worksheet, "C3:D3", bg=FAIL_RED, bold=True, value=failed_cases)
    _mc(worksheet, "E3:F3", bg=MED_BLUE, bold=True, fg=WHITE, value="Skipped")
    _sc(worksheet, "G3", bg=WHITE, bold=True, value=skipped_cases)

    headers = [
        "STT",
        "Mã test case",
        "Mô tả",
        "Kết quả mới nhất",
        "Ngày ghi nhận",
        "XML nguồn",
        "Mở chi tiết",
    ]
    for index, header in enumerate(headers, start=1):
        _sc(worksheet, f"{chr(64 + index)}4", bg=MED_BLUE, bold=True, fg=WHITE, value=header)

    for index, row in enumerate(summary_rows, start=5):
        status = row.get("result", "skipped")
        if status == "failed":
            row_bg = FAIL_RED
        elif status == "passed":
            row_bg = LIGHT_GREEN
        else:
            row_bg = WHITE

        _sc(worksheet, f"A{index}", bg=row_bg, value=index - 4)
        _sc(worksheet, f"B{index}", bg=row_bg, value=row.get("case_id", ""))
        _sc(worksheet, f"C{index}", bg=row_bg, value=_normalize_text(row.get("description", "")))
        _sc(worksheet, f"D{index}", bg=row_bg, bold=True, value=RESULT_LABELS.get(status, RESULT_LABELS["skipped"]))
        _sc(worksheet, f"E{index}", bg=row_bg, value=row.get("run_date", ""))
        _sc(worksheet, f"F{index}", bg=row_bg, value=row.get("xml_file") or "Không có")
        _sc(worksheet, f"G{index}", bg=row_bg, value="Mở sheet")

        target_sheet = row.get("sheet_name")
        if target_sheet:
            for cell_ref in (f"B{index}", f"G{index}"):
                worksheet[cell_ref].hyperlink = f"#'{target_sheet}'!A1"
                worksheet[cell_ref].style = "Hyperlink"


def _short_error(text: str, max_len: int = 200) -> str:
    if not text:
        return ""
    return _truncate_text(_extract_actionable_error(text), max_len)


def _simplify_error_text(text: str) -> str:
    normalized = _normalize_text(text).strip()
    for line in normalized.splitlines():
        cleaned = line.strip()
        if cleaned.startswith("AssertionError:"):
            return _extract_actionable_error(cleaned.removeprefix("AssertionError:").strip())
        if cleaned.startswith("E       AssertionError:"):
            return _extract_actionable_error(cleaned.removeprefix("E       AssertionError:").strip())
    return _extract_actionable_error(normalized)


def _extract_param_label(name: str) -> str:
    if "[" not in name or not name.endswith("]"):
        return ""
    return name.split("[", 1)[1][:-1]


def _module_prefix_from_classname(classname: str) -> str:
    return MODULE_PREFIX_MAP.get(classname, "AUTO")


def _base_function_name(name: str) -> str:
    return name.split("[", 1)[0].strip()


def _build_fallback_meta(classname: str, name: str, generated_ids: dict[str, int]) -> dict:
    prefix = _module_prefix_from_classname(classname)
    generated_ids[prefix] = generated_ids.get(prefix, 0) + 1
    case_id = f"{prefix}_AUTO_{generated_ids[prefix]:02d}"
    param_label = _extract_param_label(name)
    description = _default_description(_base_function_name(name))
    if param_label:
        description = f"{description} | Trường hợp: {param_label}"

    test_data = []
    if param_label:
        test_data.append(f"Tham số kiểm thử: {param_label}")

    return {
        "id": case_id,
        "description": description,
        "technique": "Automated Regression / Auto-Mapped",
        "prerequisites": [
            "Môi trường pytest và dữ liệu hiện tại cho phép thực thi testcase tự động.",
        ],
        "test_data": test_data,
        "steps": [
            {
                "detail": f"Thực thi tự động testcase `{name}` từ pytest",
                "expected": "Testcase được collect và chạy trong suite hiện tại",
            },
            {
                "detail": "Đối chiếu kết quả assert của testcase",
                "expected": "Kết quả thực tế khớp với điều kiện xác nhận trong mã test",
            },
        ],
    }


def resolve_meta_for_test(classname: str, name: str, meta_dict: dict, generated_ids: dict[str, int] | None = None) -> dict:
    direct_key = f"{classname}::{name}"
    if direct_key in meta_dict:
        return dict(meta_dict[direct_key])

    direct_canonical = canonical_token(direct_key)
    for meta_key, meta_value in meta_dict.items():
        if canonical_token(meta_key) == direct_canonical:
            return dict(meta_value)

    param_label = _extract_param_label(name)
    for meta_key, meta_value in meta_dict.items():
        meta_classname, meta_name = meta_key.split("::", 1)
        if meta_classname != classname:
            continue
        if name.startswith(meta_name + "[") or name.startswith(meta_name + " "):
            derived_meta = dict(meta_value)
            if param_label:
                derived_meta["description"] = f"{_normalize_text(meta_value.get('description', meta_name))} | Trường hợp: {param_label}"
                inherited_test_data = list(meta_value.get("test_data", []))
                inherited_test_data.append(f"Tham số kiểm thử: {param_label}")
                derived_meta["test_data"] = inherited_test_data
            return derived_meta
        if canonical_token(meta_name) == canonical_token(name):
            derived_meta = dict(meta_value)
            if param_label:
                derived_meta["description"] = f"{_normalize_text(meta_value.get('description', meta_name))} | Trường hợp: {param_label}"
                inherited_test_data = list(meta_value.get("test_data", []))
                inherited_test_data.append(f"Tham số kiểm thử: {param_label}")
                derived_meta["test_data"] = inherited_test_data
            return derived_meta

    if generated_ids is None:
        generated_ids = {}
    return _build_fallback_meta(classname, name, generated_ids)


def _severity_from_case(meta: dict, classname: str = "") -> str:
    if meta.get("severity"):
        return str(meta["severity"]).upper()
    case_id = _normalize_text(meta.get("id", "")).upper()
    if case_id.startswith(("REG", "LOG", "PWD", "DN", "DK", "GDTK")):
        return "HIGH"
    if case_id.startswith(("CRT", "CHK", "VCH", "GH")):
        return "MEDIUM"
    if any(token in classname for token in ("test_login", "test_account", "test_register", "test_runtime", "test_source")):
        return "HIGH"
    if any(token in classname for token in ("test_cart", "test_checkout")):
        return "MEDIUM"
    return "LOW"


def _module_from_case(meta: dict, classname: str = "") -> str:
    case_id = _normalize_text(meta.get("id", "")).upper()
    if case_id.startswith(("REG", "LOG", "PWD", "DN", "DK", "GDTK")):
        return "Module 1 - Quản lý tài khoản"
    if case_id.startswith(("CRT", "CHK", "VCH", "GH")):
        return "Module 2 - Giỏ hàng / Thanh toán"
    if "runtime" in classname:
        return "Runtime Audit"
    if "source" in classname:
        return "Source Audit"
    return "UI / Khác"


def _bug_category(meta: dict, error_text: str, classname: str = "") -> str:
    case_id = _normalize_text(meta.get("id", "")).upper()
    haystack = " ".join(
        [
            case_id,
            _normalize_text(meta.get("description", "")),
            _normalize_text(meta.get("technique", "")),
            _normalize_text(error_text),
            classname,
        ]
    ).lower()
    if any(token in haystack for token in ("security", "bảo mật", "mật khẩu", "password", "token", "csrf", "xss", "sql", "unauthorized", "authorization")):
        return "Security"
    if any(token in haystack for token in ("checkout", "thanh toán", "đơn hàng", "order", "voucher", "cart", "giỏ hàng", "quantity", "tổng tiền")):
        return "Business"
    if any(token in haystack for token in ("ui", "layout", "giao diện", "button", "form", "responsive")):
        return "UI"
    if "runtime" in haystack or "source" in haystack:
        return "Technical"
    return "Functional"


def _format_bug_steps(steps: list[dict]) -> str:
    if not steps:
        return (
            "1. Mở chức năng đang được kiểm thử theo mã test case.\n"
            "2. Nhập dữ liệu kiểm thử đúng như phần Test Data.\n"
            "3. Thực hiện thao tác Submit/kiểm tra theo kịch bản.\n"
            "4. Quan sát thông báo, dữ liệu lưu lại và trạng thái giao diện."
        )

    lines = []
    for index, step in enumerate(steps, start=1):
        detail = _normalize_text(step.get("detail", "")).strip()
        expected = _normalize_text(step.get("expected", "")).strip()
        if expected:
            lines.append(f"{index}. {detail}\n   Kết quả cần thấy: {expected}")
        else:
            lines.append(f"{index}. {detail}")
    return "\n".join(lines)


def _bug_expected_text(case_id: str, expected: str) -> str:
    expected = _normalize_text(expected).strip()
    if expected:
        cleaned = expected.lstrip("- ").strip()
        return f"Theo testcase {case_id}, hệ thống phải xử lý đúng yêu cầu sau:\n- {cleaned}"
    return (
        f"Theo testcase {case_id}, hệ thống phải hiển thị thông báo/điều hướng/dữ liệu đúng với đặc tả, "
        "không được lưu trạng thái sai hoặc bỏ qua validation."
    )


def _bug_actual_text(actual: str, raw: str) -> str:
    actual = _normalize_text(actual).strip()
    raw = _normalize_text(raw).strip()
    lines = []
    if actual:
        lines.append(f"Hệ thống/automation ghi nhận: {actual}")
    if raw and raw != actual:
        lines.append(f"Dấu hiệu kỹ thuật: {raw}")
    if not lines:
        lines.append("Kết quả thực tế khác với expected result của testcase nhưng chưa có message kỹ thuật chi tiết.")
    return "\n".join(lines)


def _suggest_bug_fix(meta: dict, error_text: str) -> str:
    case_id = _normalize_text(meta.get("id", "")).upper()
    description = _normalize_text(meta.get("description", ""))
    haystack = f"{case_id} {description} {_normalize_text(error_text)}".lower()

    if case_id == "PWD_05" or ("mật khẩu" in haystack and ("ngắn" in haystack or "length" in haystack or "password" in haystack)):
        return (
            'Frontend: thêm minlength="8", validate ngay dưới ô mật khẩu mới và không cho submit khi mật khẩu quá ngắn.\n'
            "Backend: trước khi lưu phải kiểm tra độ dài newPassword >= 8, đồng thời áp dụng regex có chữ hoa, chữ thường, số và ký tự đặc biệt.\n"
            "Regression: thêm test UI/API với newPassword='abc' để xác nhận server trả lỗi 400 và không cập nhật CSDL."
        )
    if "xss" in haystack or "<script" in haystack:
        return (
            "Frontend: escape dữ liệu hiển thị và không render input người dùng bằng innerHTML.\n"
            "Backend: sanitize/validate các trường địa chỉ, ghi chú trước khi lưu; encode output khi trả về frontend.\n"
            "Regression: thêm payload <script>alert('hack')</script> và xác nhận script không được thực thi."
        )
    if "sql" in haystack or "' or 1=1" in haystack:
        return (
            "Backend: dùng query parameter/ORM đúng cách, không nối chuỗi SQL trực tiếp.\n"
            "Validation: chặn hoặc escape ký tự nguy hiểm ở username/phone nếu không thuộc rule cho phép.\n"
            "Regression: thêm testcase nhập payload SQL injection và xác nhận không đăng nhập/không tạo dữ liệu sai."
        )
    if "voucher" in haystack:
        return (
            "Frontend và Backend: trim mã voucher trước khi kiểm tra.\n"
            "Backend: kiểm tra ngày hết hạn, minOrderValue, usageLimit và chống apply nhiều lần trên cùng đơn.\n"
            "Regression: test voucher hợp lệ, sai mã, hết hạn, thiếu điều kiện và dư khoảng trắng."
        )
    if "double" in haystack or "nhiều lần" in haystack or "spam" in haystack or "duplicate" in haystack:
        return (
            "Frontend: disable nút Submit ngay sau click đầu tiên và hiển thị trạng thái Đang xử lý.\n"
            "Backend: thêm idempotency key hoặc kiểm tra duplicate request để chỉ tạo một đơn hàng.\n"
            "Regression: automation spam click nút đặt hàng và xác nhận chỉ có một order được tạo."
        )
    if "quantity" in haystack or "số lượng" in haystack or "âm" in haystack:
        return (
            "Frontend: giới hạn input số lượng chỉ nhận số nguyên >= 1 và không vượt tồn kho.\n"
            "Backend: validate lại quantity trước khi tính tổng tiền, reject quantity <= 0 hoặc productId không tồn tại.\n"
            "Regression: test quantity = 0, -1, 999 và productId sai."
        )
    if any(token in haystack for token in ("token", "unauthorized", "authorization", "phân quyền", "admin")):
        return (
            "Backend: kiểm tra token/session và role ở mọi route protected, không chỉ ẩn nút trên giao diện.\n"
            "Frontend: redirect về login hoặc trang 403 khi API trả Unauthorized/Forbidden.\n"
            "Regression: dùng tài khoản Customer truy cập route Admin và xác nhận bị chặn."
        )
    return (
        "Frontend: hiển thị validation/message rõ ràng ngay tại vị trí người dùng thao tác.\n"
        "Backend: validate lại dữ liệu đầu vào trước khi lưu hoặc cập nhật trạng thái.\n"
        "Regression: giữ testcase này trong suite để xác nhận lỗi không tái diễn sau khi sửa."
    )


def build_bug_report_row(
    *,
    meta: dict,
    sheet_name: str,
    result: str,
    error_text: str,
    run_date: str,
    xml_file: str | None,
    screenshot_path: str | None = None,
    classname: str = "",
) -> dict | None:
    if result != "failed":
        return None

    normalized_steps = _normalize_steps(meta.get("steps", []))
    steps = _project_manual_steps(meta, normalized_steps)
    failed_step = steps[-1] if steps else {}
    failure_detail = _vietnamese_error_summary(error_text, _normalize_text(meta.get("description", "")), failed_step)
    test_data = _project_test_data_rows(meta, _normalize_list(meta.get("test_data", [])))
    case_id = _normalize_text(meta.get("id", sheet_name)) or sheet_name
    description = _normalize_text(meta.get("description", ""))
    category = _bug_category(meta, error_text, classname)
    title_text = f"[{category}] - {description or case_id} ({case_id})"
    raw_error = failure_detail.get("raw", _short_error(error_text))

    return {
        "sheet_name": sheet_name,
        "case_id": case_id,
        "description": description,
        "title": title_text,
        "category": category,
        "severity": _severity_from_case(meta, classname),
        "module": _module_from_case(meta, classname),
        "expected": _bug_expected_text(case_id, failure_detail.get("expected", "")),
        "actual": _bug_actual_text(failure_detail.get("actual", _short_error(error_text)), raw_error),
        "summary": failure_detail.get("summary", "Testcase không đạt."),
        "raw": raw_error,
        "retest": failure_detail.get("retest", ""),
        "test_data": "\n".join(f"- {item}" for item in test_data),
        "reproduce_steps": _format_bug_steps(steps),
        "fix_notes": _suggest_bug_fix(meta, error_text),
        "run_date": run_date,
        "xml_file": xml_file or "",
        "screenshot_path": screenshot_path or "",
    }


def add_bug_report_sheet(
    workbook,
    bug_rows: list[dict],
    title: str = "Bug_Report",
    summary_sheet_name: str = "Tong_Hop_Test_Case",
) -> None:
    worksheet = workbook.create_sheet(title=title, index=1 if workbook.sheetnames else 0)
    worksheet.freeze_panes = "A7"
    worksheet.sheet_properties.tabColor = "C00000" if bug_rows else "70AD47"

    column_widths = {
        "A": 16,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 34,
        "G": 18,
        "H": 18,
        "I": 18,
        "J": 18,
        "K": 18,
        "L": 18,
    }
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    _mc(
        worksheet,
        "A1:L1",
        bg=DARK_BLUE,
        bold=True,
        fg=WHITE,
        value="BUG REPORT - Tổng hợp lỗi phát hiện tự động từ các test case Fail",
    )
    _mc(worksheet, "A2:B2", bg=MED_BLUE, bold=True, fg=WHITE, value="Tổng bug")
    _sc(worksheet, "C2", bg=FAIL_RED if bug_rows else LIGHT_GREEN, bold=True, value=len(bug_rows))
    _mc(worksheet, "D2:E2", bg=MED_BLUE, bold=True, fg=WHITE, value="HIGH")
    _sc(worksheet, "F2", bg=FAIL_RED, bold=True, value=sum(1 for row in bug_rows if row["severity"] == "HIGH"))
    _mc(worksheet, "G2:H2", bg=MED_BLUE, bold=True, fg=WHITE, value="MEDIUM")
    _sc(worksheet, "I2", bg=YELLOW, bold=True, value=sum(1 for row in bug_rows if row["severity"] == "MEDIUM"))
    _mc(worksheet, "J2:K2", bg=MED_BLUE, bold=True, fg=WHITE, value="LOW")
    _sc(worksheet, "L2", bg=LIGHT_GRAY, bold=True, value=sum(1 for row in bug_rows if row["severity"] == "LOW"))

    _mc(
        worksheet,
        "A3:L3",
        bg=WHITE,
        value=(
            "Mục đích: dùng sheet này để ghi bug report nộp đồ án. Mỗi lỗi bên dưới có đúng cấu trúc: "
            "Description, Steps to reproduce, Expected result, Actual result, Notes/Đề xuất fix và Attachments."
        ),
    )

    if not bug_rows:
        _mc(worksheet, "A5:L7", bg=LIGHT_GREEN, bold=True, value="Không có testcase Fail trong lần chạy này, chưa phát sinh bug report.")
        return

    severity_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    bug_rows = sorted(bug_rows, key=lambda row: (severity_order.get(row["severity"], 99), row["case_id"]))

    _mc(worksheet, "A5:L5", bg=DARK_BLUE, bold=True, fg=WHITE, value="DANH SÁCH BUG PHÁT HIỆN")
    overview_headers = [
        "Bug ID",
        "Severity",
        "Nhóm lỗi",
        "Module",
        "Test Case",
        "Tiêu đề lỗi",
        "Link chi tiết",
        "Link testcase",
    ]
    overview_cols = ["A", "B", "C", "D", "E", "F", "J", "K"]
    for column, header in zip(overview_cols, overview_headers):
        _sc(worksheet, f"{column}6", bg=MED_BLUE, bold=True, fg=WHITE, value=header)

    detail_start = 8 + len(bug_rows)
    bug_ids = []
    for index, row in enumerate(bug_rows, start=1):
        excel_row = 6 + index
        bug_id = f"BUG_{index:03d}"
        bug_ids.append(bug_id)
        severity = row.get("severity", "LOW")
        row_bg = FAIL_RED if severity == "HIGH" else (YELLOW if severity == "MEDIUM" else LIGHT_GREEN)
        detail_row = detail_start + ((index - 1) * 9) + 1

        _sc(worksheet, f"A{excel_row}", bg=row_bg, bold=True, value=bug_id)
        _sc(worksheet, f"B{excel_row}", bg=row_bg, bold=True, value=severity)
        _sc(worksheet, f"C{excel_row}", bg=row_bg, value=row.get("category", "Functional"))
        _sc(worksheet, f"D{excel_row}", bg=row_bg, value=row.get("module", ""))
        _sc(worksheet, f"E{excel_row}", bg=row_bg, bold=True, value=row.get("case_id", ""))
        _mc(worksheet, f"F{excel_row}:I{excel_row}", bg=row_bg, bold=True, value=row.get("title", row.get("summary", "")))
        _sc(worksheet, f"J{excel_row}", bg=row_bg, bold=True, value="Mở bug")
        _sc(worksheet, f"K{excel_row}", bg=row_bg, bold=True, value="Mở testcase")
        _sc(worksheet, f"L{excel_row}", bg=row_bg, value=row.get("xml_file", ""))

        worksheet[f"J{excel_row}"].hyperlink = f"#'{title}'!A{detail_row}"
        worksheet[f"J{excel_row}"].style = "Hyperlink"
        target_sheet = row.get("sheet_name")
        if target_sheet:
            for cell_ref in (f"E{excel_row}", f"K{excel_row}"):
                worksheet[cell_ref].hyperlink = f"#'{target_sheet}'!A1"
                worksheet[cell_ref].style = "Hyperlink"

    _mc(worksheet, f"A{detail_start}:L{detail_start}", bg=DARK_BLUE, bold=True, fg=WHITE, value="CHI TIẾT BUG REPORT")

    for index, row in enumerate(bug_rows, start=1):
        bug_id = bug_ids[index - 1]
        block_row = detail_start + ((index - 1) * 9) + 1
        severity = row.get("severity", "LOW")
        header_bg = "C00000" if severity == "HIGH" else ("FFC000" if severity == "MEDIUM" else "70AD47")
        header_fg = WHITE if severity in {"HIGH", "LOW"} else BLACK
        screenshot_path = row.get("screenshot_path")
        target_sheet = row.get("sheet_name")
        attachment_text = "Chưa có ảnh lỗi."
        if screenshot_path:
            attachment_text = f"Ảnh lỗi: {Path(screenshot_path).name}"
        if row.get("xml_file"):
            attachment_text += f"\nXML nguồn: {row.get('xml_file')}"

        _mc(
            worksheet,
            f"A{block_row}:L{block_row}",
            bg=header_bg,
            bold=True,
            fg=header_fg,
            value=f"{bug_id} | {row.get('title', row.get('summary', ''))} | Severity: {severity} | Test Case: {row.get('case_id', '')}",
        )

        description_block = (
            f"Mô tả: {row.get('summary', '')}\n"
            f"Module: {row.get('module', '')}\n"
            f"Ngày phát hiện: {row.get('run_date', '')}\n"
            f"Dữ liệu kiểm thử:\n{row.get('test_data', '-') or '-'}"
        )
        bug_sections = [
            ("Description", description_block, WHITE, 70),
            ("Steps to reproduce", row.get("reproduce_steps", ""), WHITE, 115),
            ("Expected result", row.get("expected", ""), LIGHT_GREEN, 70),
            ("Actual result", row.get("actual", ""), FAIL_RED, 80),
            ("Notes (Đề xuất fix)", row.get("fix_notes", ""), YELLOW, 95),
        ]

        for offset, (label, value, bg, height) in enumerate(bug_sections, start=1):
            current_row = block_row + offset
            worksheet.row_dimensions[current_row].height = height
            _mc(worksheet, f"A{current_row}:B{current_row}", bg=MED_BLUE, bold=True, fg=WHITE, value=label)
            _mc(worksheet, f"C{current_row}:L{current_row}", bg=bg, value=value)

        attachment_row = block_row + 6
        worksheet.row_dimensions[attachment_row].height = 48
        _mc(worksheet, f"A{attachment_row}:B{attachment_row}", bg=MED_BLUE, bold=True, fg=WHITE, value="Attachments")
        _mc(worksheet, f"C{attachment_row}:F{attachment_row}", bg=WHITE, value=attachment_text)
        _mc(worksheet, f"G{attachment_row}:I{attachment_row}", bg=WHITE, value="Mở ảnh lỗi" if screenshot_path else "Không có ảnh lỗi")
        _mc(worksheet, f"J{attachment_row}:L{attachment_row}", bg=WHITE, value="Mở testcase chi tiết")

        if screenshot_path:
            try:
                worksheet[f"G{attachment_row}"].hyperlink = Path(screenshot_path).resolve().as_uri()
                worksheet[f"G{attachment_row}"].style = "Hyperlink"
            except ValueError:
                pass
        if target_sheet:
            worksheet[f"J{attachment_row}"].hyperlink = f"#'{target_sheet}'!A1"
            worksheet[f"J{attachment_row}"].style = "Hyperlink"

        spacer_row = block_row + 7
        _mc(worksheet, f"A{spacer_row}:L{spacer_row}", bg=LIGHT_GRAY, value="")

    final_row = detail_start + (len(bug_rows) * 9) + 2
    _mc(worksheet, f"A{final_row}:L{final_row}", bg=DARK_BLUE, bold=True, fg=WHITE, value="Gợi ý xử lý bug")
    _mc(
        worksheet,
        f"A{final_row + 1}:L{final_row + 3}",
        bg=WHITE,
        value=(
            "Quy trình đề xuất: 1) Mở bug detail. 2) Làm lại Steps to reproduce bằng tay trên web thật. "
            "3) Đối chiếu Expected result và Actual result. 4) Nếu lỗi còn tồn tại, tạo bug với trạng thái New. "
            "5) Sau khi dev sửa, chạy lại testcase và chuyển trạng thái Fixed -> Retest -> Closed."
        ),
    )


def _parse_xml(xml_path: Path) -> list[tuple[str, str, str, str]]:
    tree = ET.parse(xml_path)
    root = tree.getroot()
    records: list[tuple[str, str, str, str]] = []

    for testcase in root.iter("testcase"):
        classname = testcase.get("classname", "")
        name = testcase.get("name", "")

        failure = testcase.find("failure")
        error = testcase.find("error")
        skipped = testcase.find("skipped")

        if failure is not None:
            result = "failed"
            error_text = _simplify_error_text(failure.get("message", "") or (failure.text or ""))
        elif error is not None:
            result = "failed"
            error_text = _simplify_error_text(error.get("message", "") or (error.text or ""))
        elif skipped is not None:
            result = "skipped"
            error_text = ""
        else:
            result = "passed"
            error_text = ""

        records.append((classname, name, result, error_text))

    return records


def _lookup_meta(classname: str, name: str, meta_dict: dict) -> dict | None:
    return resolve_meta_for_test(classname, name, meta_dict, {})


def load_artifacts_for_xml(xml_path: Path) -> dict:
    xml_path = Path(xml_path)
    timestamp = xml_path.stem.removeprefix("results_")
    artifacts_index = xml_path.parent / f"artifacts_{timestamp}.json"
    if not artifacts_index.exists():
        return {}
    try:
        payload = json.loads(artifacts_index.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    return payload.get("screenshots", {})


def generate_report(xml_path, output_path=None, run_date=None):
    from catalog.test_case_meta import TEST_CASE_META

    if run_date is None:
        run_date = datetime.now().strftime("%d/%m/%Y")

    xml_path = Path(xml_path)
    if output_path is None:
        if xml_path.stem.startswith("results_"):
            timestamp = xml_path.stem.removeprefix("results_")
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = xml_path.parent / f"report_excel_{timestamp}.xlsx"
    output_path = Path(output_path)

    records = _parse_xml(xml_path)
    screenshot_artifacts = load_artifacts_for_xml(xml_path)
    workbook = Workbook()
    workbook.remove(workbook.active)

    seen_ids: dict[str, int] = {}
    generated_meta_ids: dict[str, int] = {}
    sheet_count = 0
    summary_rows: list[dict] = []
    bug_rows: list[dict] = []

    for classname, name, result, error_text in records:
        meta = resolve_meta_for_test(classname, name, TEST_CASE_META, generated_meta_ids)
        artifact = screenshot_artifacts.get(f"{classname}::{name}", {})
        screenshot_path = artifact.get("path")

        base_id = meta["id"]
        if base_id in seen_ids:
            seen_ids[base_id] += 1
            sheet_name = f"{base_id}_{seen_ids[base_id]}"
        else:
            seen_ids[base_id] = 0
            sheet_name = base_id

        worksheet = workbook.create_sheet(title=sheet_name)
        sheet_meta = {**meta, "id": sheet_name}
        build_sheet(
            worksheet,
            sheet_meta,
            result,
            error_text,
            run_date,
            summary_sheet_name="Tong_Hop_Test_Case",
            screenshot_path=screenshot_path,
        )
        bug_row = build_bug_report_row(
            meta=sheet_meta,
            sheet_name=sheet_name,
            result=result,
            error_text=error_text,
            run_date=run_date,
            xml_file=xml_path.name,
            screenshot_path=screenshot_path,
            classname=classname,
        )
        if bug_row:
            bug_rows.append(bug_row)
        summary_rows.append(
            {
                "sheet_name": sheet_name,
                "case_id": sheet_name,
                "description": sheet_meta["description"],
                "result": result,
                "run_date": run_date,
                "xml_file": xml_path.name,
            }
        )
        sheet_count += 1

    if sheet_count == 0:
        worksheet = workbook.create_sheet(title="Khong_Co_Test_Map")
        worksheet["A1"] = (
            "Không tìm thấy test case nào khớp trong catalog/test_case_meta.py. "
            "Hãy kiểm tra lại định dạng key theo mẫu 'classname::test_name'."
        )
    else:
        add_summary_sheet(workbook, summary_rows, title="Tong_Hop_Test_Case")
        add_bug_report_sheet(workbook, bug_rows, title="Bug_Report", summary_sheet_name="Tong_Hop_Test_Case")

    workbook.save(output_path)
    print(f"[Excel Report] {sheet_count} sheet(s) -> {output_path}")
    return output_path


def _latest_xml(reports_dir="reports"):
    xml_files = sorted(Path(reports_dir).glob("results_*.xml"), key=lambda item: item.stat().st_mtime)
    return xml_files[-1] if xml_files else None


if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    if len(sys.argv) >= 2:
        xml_file = Path(sys.argv[1])
    else:
        xml_file = _latest_xml()
        if xml_file is None:
            print("[Excel Report] Không tìm thấy file XML trong reports/. Hãy chạy pytest trước.")
            sys.exit(1)
        print(f"[Excel Report] Sử dụng XML mới nhất: {xml_file}")

    output_file = generate_report(xml_file)
    print(f"[Excel Report] Đã lưu: {output_file}")
