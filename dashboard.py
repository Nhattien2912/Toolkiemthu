from __future__ import annotations

import json
import os
import queue
import subprocess
import sys
import threading
import webbrowser
from datetime import datetime
from openpyxl import load_workbook

try:
    import psutil
except ImportError:  # pragma: no cover - dependency fallback for environments not yet updated
    psutil = None

from flask import Flask, Response, abort, redirect, render_template, request, send_from_directory, url_for

from utils.cumulative_reporter import build_latest_results, generate_cumulative_report, get_coverage_summary
from utils.excel_reporter import SUMMARY_SHEET_NAMES
from utils.pdf_reporter import generate_pdf_report
from utils.project_doc_data import (
    BASE37_CASE_CODES,
    CUSTOM_EXTRA_TEST_CASE_IDS,
    EXCEL62_TEST_NODEIDS,
    EXTENDED80_CASE_CODES,
    PROJECT_WORD_DATA,
    RISK_TEST_CASE_IDS,
    TEST_CASE_SUBGROUPS,
)
from utils.report_parser import (
    PROJECT_ROOT,
    REPORTS_DIR,
    build_case_sheet,
    compare_runs,
    format_duration,
    get_all_runs,
    get_automated_insights,
    get_bug_tracker_rows,
    get_critical_failures,
    get_historical_data,
    get_module_trend_data,
)
from utils.word_reporter import generate_word_report


app = Flask(__name__)
app.config["JSON_AS_ASCII"] = False
app.json.ensure_ascii = False

LAST_LOG_FILE = REPORTS_DIR / "_last_dashboard_run.log"
LAST_STATE_FILE = REPORTS_DIR / "_last_dashboard_run.json"
_run_lock = threading.Lock()
_log_queue: queue.Queue[str] = queue.Queue()
_active_process: subprocess.Popen[str] | None = None
_run_in_progress = False
_run_paused = False
_pytest_command_cache: list[str] | None = None
_pytest_command_probed = False

REAL_FULL_ARGS = [
    "tests/test_login.py",
    "tests/test_register_module.py",
    "tests/test_account_module.py",
    "tests/test_cart_checkout_module.py",
    "tests/test_runtime_audit.py",
    "tests/test_source_audit.py",
    "tests/test_source_conformance.py",
    "tests/test_source_deep_conformance.py",
    "tests/test_custom_extra_cases.py",
    "tests/test_risk_cases.py",
    "tests/test_ui.py",
    "tests/test_search.py",
    "tests/test_pages.py",
    "tests/test_home_filters.py",
    "tests/test_home_products.py",
    "tests/test_catalog_integrity.py",
    "-v",
]

TEST_SUITES = {
    "full": {
        "label": "Full Regression",
        "description": "Chay toan bo test automation hien co.",
        "method": "Regression / End-to-End",
        "args": REAL_FULL_ARGS,
    },
    "project-max": {
        "label": "Run Max - Toan bo test that",
        "description": "Chay toan bo test that dang co trong repo, khong gioi han trong 100 case cua pham vi do an va khong gom demo_failures.",
        "method": "Maximum Coverage / Real Test Repo",
        "args": REAL_FULL_ARGS,
    },
    "project-all": {
        "label": "Run All - Theo Do An",
        "description": "Chay toan bo pham vi do an: Module 1, Module 2, strict runtime audit, strict source audit va source conformance trong mot lan.",
        "method": "Project Scope / Thesis Run",
        "args": [
            "tests/test_login.py",
            "tests/test_register_module.py",
            "tests/test_account_module.py",
            "tests/test_cart_checkout_module.py",
            "tests/test_runtime_audit.py",
            "tests/test_source_audit.py",
            "tests/test_source_conformance.py",
            "tests/test_risk_cases.py",
            "-v",
        ],
    },
    "tc-62-suite": {
        "label": "Run 117 Test Care Suite",
        "description": "Alias cu, chay toan bo 37 TC nen + 80 TC mo rong khong trung.",
        "method": "Report-Aligned / Full 117 Cases",
        "args": [
            "tests/test_tc62_suite.py",
            "-v",
        ],
    },
    "excel-62-suite": {
        "label": "Run 62 Test Chi Tiết",
        "description": "Chạy đúng danh sách testcase lấy từ file Excel Danh sách 62 Test Case Chi Tiết (1).xlsx, dùng nodeid cụ thể để không chạy lẫn bộ 117.",
        "method": "Excel-Aligned / Detailed 62 Test Case Workbook",
        "args": [
            *EXCEL62_TEST_NODEIDS,
            "-v",
        ],
    },
    "tc-117-suite": {
        "label": "Run 117 Test Care Suite",
        "description": "Chay toan bo test case do an: 37 TC nen va 80 TC mo rong khong trung ma.",
        "method": "Report-Aligned / Full 117 Cases",
        "args": [
            "tests/test_tc62_suite.py",
            "-v",
        ],
    },
    "tc-37-base-suite": {
        "label": "Run 37 Test Case Goc",
        "description": "Chay rieng 37 test case nen dung theo bang goc ban dua.",
        "method": "Report-Aligned / Base 37 Cases",
        "args": [
            "tests/test_tc62_suite.py",
            "-k",
            " or ".join(BASE37_CASE_CODES),
            "-v",
        ],
    },
    "tc-80-suite": {
        "label": "Run 80 Test Case Mo Rong",
        "description": "Chay rieng 80 test case mo rong, khong trung voi 37 test case goc trong bang ban dua.",
        "method": "Report-Aligned / Non-Duplicate Extension",
        "args": [
            "tests/test_tc62_suite.py",
            "-k",
            " or ".join(EXTENDED80_CASE_CODES),
            "-v",
        ],
    },
    "tien-test-cases": {
        "label": "Test Case Của Tiến",
        "description": "Chạy riêng 28 test case của Tiến từ REG_13 đến VCH_08, tập trung validation, cart, checkout và voucher.",
        "method": "Focused Suite / Test Case Của Tiến",
        "args": [
            "tests/test_tien_test_cases.py",
            "-v",
        ],
    },
    "custom-extra-cases": {
        "label": "Test Case Mới 66-85",
        "description": "Chạy riêng 24 testcase mới bổ sung: REG_18/19/20/22, CRT_18-22 và CHK_23-27. Các mã bị trùng được tách bằng unique_id nội bộ để chạy đủ từng dòng.",
        "method": "Focused Suite / Extra Edge Cases 66-85",
        "args": [
            "tests/test_custom_extra_cases.py",
            "-k",
            " or ".join(CUSTOM_EXTRA_TEST_CASE_IDS),
            "-v",
        ],
    },
    "risk-cases": {
        "label": "Run 6 Case Rủi Ro",
        "description": "Chạy riêng 6 testcase rủi ro cao: logout, voucher trim, notes cực dài, XSS checkout, chống double-click và payload quantity âm.",
        "method": "Focused Suite / Security and Critical Checkout Risk",
        "args": [
            "tests/test_risk_cases.py",
            "-k",
            " or ".join(RISK_TEST_CASE_IDS),
            "-v",
        ],
    },
    "reg-12-suite": {
        "label": "Run 12 Test Care Dau Tien",
        "description": "Chay rieng 12 test case dau tien cua nhom Dang ky: REG_01 den REG_12, theo dung bang test case trong bao cao.",
        "method": "Register Requirement / Detailed Validation",
        "args": [
            "tests/test_tc62_suite.py",
            "-k",
            " or ".join([f"REG_{index:02d}" for index in range(1, 13)]),
            "-v",
        ],
    },
    "module-1": {
        "label": "Module 1 - Account Management",
        "description": "Dang ky, dang nhap, doi mat khau, kiem tra form account va doi chieu loi forgot/reset.",
        "method": "Black-box / Validation / UI",
        "args": [
            "tests/test_login.py",
            "tests/test_register_module.py",
            "tests/test_account_module.py",
            "-v",
        ],
    },
    "module-2": {
        "label": "Module 2 - Cart and Checkout",
        "description": "Gio hang, voucher va thanh toan theo pham vi do an.",
        "method": "Business Flow / Regression",
        "args": [
            "tests/test_cart_checkout_module.py",
            "-v",
        ],
    },
    "uc-01-su": {
        "label": "UC_01_SU - Register",
        "description": "Kiem thu ky chuc nang dang ky tai khoan.",
        "method": "Black-box / Equivalence / Boundary",
        "args": [
            "tests/test_register_module.py",
            "-v",
        ],
    },
    "uc-02-si": {
        "label": "UC_02_SI - Login",
        "description": "Kiem thu ky chuc nang dang nhap, validation va session login.",
        "method": "Black-box / Validation / Session",
        "args": [
            "tests/test_login.py",
            "tests/test_account_module.py",
            "-k",
            "login",
            "-v",
        ],
    },
    "uc-03-fp": {
        "label": "UC_03_FP - Forgot Password",
        "description": "Kiem tra diem vao va dieu huong cua chuc nang quen mat khau.",
        "method": "Navigation / Access Check",
        "args": [
            "tests/test_account_module.py",
            "-k",
            "forgot",
            "-v",
        ],
    },
    "uc-03-cp": {
        "label": "UC_03_CP - Change Password",
        "description": "Kiem thu tab doi mat khau va bo cuc account profile.",
        "method": "Black-box / Account Security",
        "args": [
            "tests/test_account_module.py",
            "-k",
            "change_password_tab or profile_page_form_and_change_password_tab",
            "-v",
        ],
    },
    "uc-04-add-to-cart": {
        "label": "UC_04_AD_TO_CART - Add To Cart",
        "description": "Kiem thu gio hang da seed du lieu, headers, voucher va tinh toan hien thi.",
        "method": "Black-box / Cart Flow",
        "args": [
            "tests/test_cart_checkout_module.py",
            "-k",
            "cart_page",
            "-v",
        ],
    },
    "uc-05-checkout": {
        "label": "UC_05_CHECKOUT - Checkout",
        "description": "Kiem thu form checkout, payment controls va submit validation.",
        "method": "Black-box / Checkout Flow",
        "args": [
            "tests/test_cart_checkout_module.py",
            "-k",
            "checkout",
            "-v",
        ],
    },
    "auth": {
        "label": "Account and Auth",
        "description": "Dang nhap, dang ky, validation va account profile.",
        "method": "Functional / Validation",
        "args": [
            "tests/test_login.py",
            "tests/test_register_module.py",
            "tests/test_account_module.py",
        ],
    },
    "cart": {
        "label": "Cart and Checkout",
        "description": "Gio hang, voucher, checkout va summary.",
        "method": "Business Flow / Regression",
        "args": [
            "tests/test_cart_checkout_module.py",
            "tests/test_pages.py",
        ],
    },
    "ui": {
        "label": "UI and Navigation",
        "description": "Homepage, search, page layout, broken image/link level co ban.",
        "method": "UI / Usability / Smoke",
        "args": [
            "tests/test_ui.py",
            "tests/test_search.py",
            "tests/test_pages.py",
            "tests/test_home_filters.py",
            "tests/test_home_products.py",
        ],
    },
    "catalog": {
        "label": "Catalog Integrity",
        "description": "Kiem tra catalog metadata va ID mapping.",
        "method": "Meta Validation",
        "args": ["tests/test_catalog_integrity.py"],
    },
    "strict-audit": {
        "label": "Strict Runtime Audit",
        "description": "Bat loi console, broken resource, broken deep-link va runtime regression.",
        "method": "Quality Gate / Deployment Audit",
        "args": ["tests/test_runtime_audit.py", "-v"],
    },
    "strict-source-audit": {
        "label": "Strict Source Audit",
        "description": "Doi chieu source code goc de bat loi route, validation, contract va authorization gap.",
        "method": "Source Audit / Contract Review",
        "args": ["tests/test_source_audit.py", "-v"],
    },
    "source-conformance": {
        "label": "Source Conformance",
        "description": "Kiem tra cac route, control, schema va logic chuan hien co trong source.",
        "method": "Source Review / Conformance",
        "args": ["tests/test_source_conformance.py", "tests/test_source_deep_conformance.py", "-v"],
    },
    "fail-demo": {
        "label": "Failure Demo",
        "description": "Bo test co y fail de demo report va failure trace.",
        "method": "Negative / Demo",
        "args": ["tests/demo_failures.py", "-v"],
    },
}

for subgroup in TEST_CASE_SUBGROUPS:
    TEST_SUITES[subgroup["suite_id"]] = {
        "label": subgroup["title"],
        "description": f"{subgroup['focus']} Ma TC: {', '.join(subgroup['codes'])}.",
        "method": f"TC80 Subset / {subgroup['technique']}",
        "args": [
            "tests/test_tc62_suite.py",
            "-k",
            " or ".join(subgroup["codes"]),
            "-v",
        ],
    }


def _pytest_command_candidates() -> list[list[str]]:
    candidates: list[list[str]] = [[sys.executable, "-m", "pytest"]]
    if os.name == "nt":
        candidates.extend(
            [
                ["python", "-m", "pytest"],
                ["py", "-3.12", "-m", "pytest"],
                ["py", "-m", "pytest"],
            ]
        )
    else:
        candidates.append(["python", "-m", "pytest"])
    return candidates


def _resolve_pytest_command() -> list[str] | None:
    global _pytest_command_cache, _pytest_command_probed

    if _pytest_command_probed:
        return _pytest_command_cache

    for command in _pytest_command_candidates():
        try:
            probe = subprocess.run(
                [*command, "--version"],
                cwd=PROJECT_ROOT,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=15,
                check=False,
            )
        except (FileNotFoundError, subprocess.TimeoutExpired):
            continue

        if probe.returncode == 0:
            _pytest_command_cache = command
            _pytest_command_probed = True
            return command

    _pytest_command_cache = None
    _pytest_command_probed = True
    return None


def _command_preview_prefix() -> str:
    command = _resolve_pytest_command()
    if command:
        return " ".join(command)
    return "python -m pytest"


def _build_suite_command_preview(suite: dict) -> str:
    preview_parts = [_command_preview_prefix(), *suite.get("args", [])]
    return " ".join(part for part in preview_parts if part).strip()


def _decorate_suite(suite_id: str, suite: dict) -> dict:
    decorated = {"id": suite_id, **suite}
    decorated["command_preview"] = _build_suite_command_preview(suite)
    return decorated


def _get_suite_choices() -> list[dict]:
    return [_decorate_suite(suite_id, suite) for suite_id, suite in TEST_SUITES.items()]


def _get_suite(suite_id: str | None) -> tuple[str, dict]:
    normalized = suite_id or "full"
    if normalized not in TEST_SUITES:
        normalized = "full"
    return normalized, TEST_SUITES[normalized]


def _load_run_state() -> dict | None:
    if not LAST_STATE_FILE.exists():
        return None
    try:
        return json.loads(LAST_STATE_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None


def _save_run_state(state: dict) -> None:
    LAST_STATE_FILE.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _merge_run_state(**updates) -> dict | None:
    current = _load_run_state()
    if current is None:
        return None
    current.update(updates)
    _save_run_state(current)
    return current


def _is_pytest_process(pid: int | None) -> bool:
    if psutil is None or pid is None:
        return False
    try:
        process = psutil.Process(pid)
        commandline = " ".join(process.cmdline()).lower()
    except psutil.Error:
        return False
    return "pytest" in commandline


def _get_live_run_pid() -> int | None:
    if _active_process is not None and _active_process.poll() is None:
        return _active_process.pid

    state = _load_run_state()
    if not state:
        return None

    pid = state.get("pid")
    if isinstance(pid, int) and _is_pytest_process(pid):
        return pid
    return None


def _refresh_live_run_state() -> dict | None:
    state = _load_run_state()
    if not state:
        return None

    live_pid = _get_live_run_pid()
    if live_pid is not None:
        if state.get("pid") != live_pid:
            state["pid"] = live_pid
            _save_run_state(state)
        return state

    if state.get("status") in {"running", "paused"} or state.get("paused"):
        state.update(
            {
                "status": "interrupted",
                "paused": False,
                "pid": None,
                "completed_at": datetime.now().astimezone().isoformat(timespec="seconds"),
                "message": "Khong con tien trinh pytest dang chay. Trang thai cu da duoc reset.",
            }
        )
        _save_run_state(state)
    return state


def _iter_process_tree(root_pid: int):
    if psutil is None:
        return []
    try:
        root_process = psutil.Process(root_pid)
    except psutil.Error:
        return []
    return [*root_process.children(recursive=True), root_process]


def _suspend_active_process_tree() -> bool:
    global _run_paused

    root_pid = _get_live_run_pid()
    if psutil is None or root_pid is None:
        return False

    processes = _iter_process_tree(root_pid)
    if not processes:
        return False

    for process in processes:
        try:
            process.suspend()
        except psutil.Error:
            continue

    _run_paused = True
    _merge_run_state(
        paused=True,
        status="paused",
        message="Suite dang tam dung. Ban co the tiep tuc tu giao dien project-doc/dashboard.",
    )
    return True


def _resume_active_process_tree() -> bool:
    global _run_paused

    root_pid = _get_live_run_pid()
    if psutil is None or root_pid is None:
        return False

    processes = _iter_process_tree(root_pid)
    if not processes:
        return False

    for process in reversed(processes):
        try:
            process.resume()
        except psutil.Error:
            continue

    _run_paused = False
    current_state = _load_run_state() or {}
    _merge_run_state(
        paused=False,
        status="running",
        message=f"Suite '{current_state.get('suite_label', 'current run')}' tiep tuc chay.",
    )
    return True


def _get_run_by_filename(xml_filename: str | None) -> dict | None:
    if not xml_filename:
        return None
    for run in get_all_runs():
        if run["xml_filename"] == xml_filename:
            return run
    return None


def _pick_generated_run(before_xml_files: set[str]) -> dict | None:
    new_xml_files = sorted(
        [path for path in REPORTS_DIR.glob("results_*.xml") if path.name not in before_xml_files],
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    if not new_xml_files:
        return None
    return _get_run_by_filename(new_xml_files[0].name)


def _safe_report_path(filename: str):
    if "/" in filename or "\\" in filename:
        abort(400)
    if not filename.startswith("report_") or not filename.endswith(".html"):
        abort(404)
    candidate = REPORTS_DIR / filename
    if not candidate.exists():
        abort(404)
    if candidate.resolve().parent != REPORTS_DIR.resolve():
        abort(403)
    return candidate


def _safe_excel_path(filename: str):
    if "/" in filename or "\\" in filename:
        abort(400)
    if not filename.endswith(".xlsx"):
        abort(404)
    candidate = REPORTS_DIR / filename
    if not candidate.exists():
        abort(404)
    if candidate.resolve().parent != REPORTS_DIR.resolve():
        abort(403)
    return candidate


def _safe_report_artifact_path(artifact_path: str) -> Path:
    candidate = (REPORTS_DIR / artifact_path).resolve()
    reports_root = REPORTS_DIR.resolve()
    if reports_root not in candidate.parents and candidate != reports_root:
        abort(403)
    if not candidate.exists() or not candidate.is_file():
        abort(404)
    return candidate


def _find_run(xml_filename: str | None):
    runs = get_all_runs()
    if not runs:
        return None
    if not xml_filename:
        return runs[0]
    for run in runs:
        if run["xml_filename"] == xml_filename:
            return run
    return None


def _format_file_size(size_bytes: int) -> str:
    if size_bytes < 1024:
        return f"{size_bytes} B"
    if size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes / (1024 * 1024):.2f} MB"


def _get_excel_reports() -> list[dict]:
    excel_reports = []
    for path in REPORTS_DIR.glob("*.xlsx"):
        stats = path.stat()
        created_at = datetime.fromtimestamp(stats.st_ctime).astimezone()
        modified_at = datetime.fromtimestamp(stats.st_mtime).astimezone()
        excel_metrics = _extract_excel_metrics(path)
        xml_filename = None
        html_filename = None

        if path.stem.startswith("report_excel_"):
            timestamp_token = path.stem.removeprefix("report_excel_")
            derived_xml = f"results_{timestamp_token}.xml"
            derived_html = f"report_{timestamp_token}.html"
            if (REPORTS_DIR / derived_xml).exists():
                xml_filename = derived_xml
            if (REPORTS_DIR / derived_html).exists():
                html_filename = derived_html

        excel_reports.append(
            {
                "filename": path.name,
                "created_at_epoch": stats.st_ctime,
                "created_at_display": created_at.strftime("%d/%m/%Y %H:%M:%S"),
                "modified_at_display": modified_at.strftime("%d/%m/%Y %H:%M:%S"),
                "size_display": _format_file_size(stats.st_size),
                "total_cases": excel_metrics["total_cases"],
                "passed_cases": excel_metrics["passed_cases"],
                "failed_cases": excel_metrics["failed_cases"],
                "skipped_cases": excel_metrics["skipped_cases"],
                "failed_sheet_names": excel_metrics["failed_sheet_names"],
                "status": excel_metrics["status"],
                "status_label": excel_metrics["status_label"],
                "xml_filename": xml_filename,
                "html_filename": html_filename,
            }
        )

    excel_reports.sort(key=lambda item: item["created_at_epoch"], reverse=True)
    return excel_reports


def _extract_excel_metrics(path) -> dict:
    metrics = {
        "total_cases": 0,
        "passed_cases": 0,
        "failed_cases": 0,
        "skipped_cases": 0,
        "failed_sheet_names": [],
        "status": "unknown",
        "status_label": "Không xác định",
    }

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        metrics["status_label"] = "Không đọc được file"
        return metrics

    try:
        for sheet_name in workbook.sheetnames:
            if sheet_name in {"Khong_Co_Test_Map", "No_Mapped_Tests"} or sheet_name in SUMMARY_SHEET_NAMES:
                continue

            worksheet = workbook[sheet_name]
            metrics["total_cases"] += 1
            raw_status = str(worksheet["K6"].value or "").strip().lower()

            if "không đạt" in raw_status or "fail" in raw_status:
                metrics["failed_cases"] += 1
                metrics["failed_sheet_names"].append(sheet_name)
            elif "chưa thực hiện" in raw_status or "not executed" in raw_status or "skipped" in raw_status:
                metrics["skipped_cases"] += 1
            elif "đạt" in raw_status or "pass" in raw_status:
                metrics["passed_cases"] += 1
    finally:
        workbook.close()

    if metrics["failed_cases"] > 0:
        metrics["status"] = "failed"
        metrics["status_label"] = "Có lỗi"
    elif metrics["total_cases"] > 0 and metrics["passed_cases"] == metrics["total_cases"]:
        metrics["status"] = "passed"
        metrics["status_label"] = "Đạt hết"
    elif metrics["total_cases"] > 0:
        metrics["status"] = "partial"
        metrics["status_label"] = "Có case chưa chạy"

    return metrics


def _build_reports_context(selected_run: dict, all_runs: list[dict], last_run_state: dict | None = None) -> dict:
    historical = get_historical_data(all_runs, n=7)
    module_trend = get_module_trend_data(all_runs, n=10)
    critical_failures = get_critical_failures(selected_run)
    slowdown = selected_run["avg_time_case_seconds"] > 20
    integrity_notice = None
    if last_run_state and last_run_state.get("status") not in {"passed", "test_failures"}:
        integrity_notice = (
            "Lan chay gan nhat khong tao duoc report hop le. Trang nay dang hien thi report hop le moi nhat, "
            "khong phai ket qua cua lan run vua loi."
        )
    elif last_run_state and last_run_state.get("linked_xml_filename") and (
        last_run_state.get("linked_xml_filename") != selected_run["xml_filename"]
    ):
        integrity_notice = (
            f"Ban dang xem run cu. Lan dashboard-triggered gan nhat la {last_run_state.get('linked_xml_filename')}."
        )
    return {
        "selected_run": selected_run,
        "critical_failures": critical_failures,
        "historical": historical,
        "module_trend": module_trend,
        "automated_insights": get_automated_insights(selected_run, all_runs),
        "all_runs": all_runs,
        "slowdown": slowdown,
        "resource_cpu": "Nominal",
        "resource_db": "Nominal",
        "resource_queue": "Clear",
        "last_run_state": last_run_state,
        "integrity_notice": integrity_notice,
    }


def _resolve_selected_case_index(run: dict, selected_case_raw: str | None) -> int:
    testcases = run["testcases"]
    if not testcases:
        return 0

    if selected_case_raw is not None:
        try:
            selected_index = int(selected_case_raw)
            if 0 <= selected_index < len(testcases):
                return selected_index
        except ValueError:
            pass

    for index, testcase in enumerate(testcases):
        if testcase["status"] in {"failed", "error"}:
            return index
    return 0


def _get_project_doc_context() -> dict:
    from catalog.test_case_meta import TEST_CASE_META as _meta

    quick_suite_ids = [
        "reg-12-suite",
        "excel-62-suite",
        "tien-test-cases",
        "custom-extra-cases",
        "risk-cases",
        "tc-37-base-suite",
        "tc-80-suite",
        "tc-117-suite",
        "project-all",
        "project-max",
        "module-1",
        "uc-01-su",
        "uc-02-si",
        "uc-03-cp",
        "module-2",
        "uc-04-add-to-cart",
        "uc-05-checkout",
        "strict-audit",
        "strict-source-audit",
        "source-conformance",
    ]
    quick_suites = [_decorate_suite(suite_id, TEST_SUITES[suite_id]) for suite_id in quick_suite_ids if suite_id in TEST_SUITES]
    suite_lookup = {suite["id"]: suite for suite in _get_suite_choices()}
    all_runs = get_all_runs()
    latest_run = all_runs[0] if all_runs else None
    _latest = build_latest_results(REPORTS_DIR)
    cumulative_coverage = get_coverage_summary(_latest, _meta)
    last_run_state = _refresh_live_run_state()
    live_pid = _get_live_run_pid()
    return {
        "project_doc": PROJECT_WORD_DATA,
        "quick_suites": quick_suites,
        "suite_lookup": suite_lookup,
        "last_run_state": last_run_state,
        "suite_running": live_pid is not None,
        "suite_paused": live_pid is not None and bool((last_run_state or {}).get("paused")),
        "selected_suite_id": (last_run_state or {}).get("suite_id", "module-1"),
        "latest_available_run": latest_run,
        "cumulative_coverage": cumulative_coverage,
    }


def _broadcast_log_line(line: str) -> None:
    _log_queue.put(line)


def _open_browser_on_start() -> None:
    if os.environ.get("DASHBOARD_NO_BROWSER") == "1":
        return
    threading.Timer(1.0, lambda: webbrowser.open("http://127.0.0.1:5050/dashboard")).start()


def _run_pytest_worker(suite_id: str) -> None:
    global _active_process, _run_in_progress, _run_paused

    suite_id, suite = _get_suite(suite_id)
    command_prefix = _resolve_pytest_command()
    before_xml_files = {path.name for path in REPORTS_DIR.glob("results_*.xml")}
    started_at = datetime.now().astimezone().isoformat(timespec="seconds")
    _run_paused = False
    try:
        with LAST_LOG_FILE.open("w", encoding="utf-8") as log_file:
            if command_prefix is None:
                missing_message = (
                    "[dashboard] Could not find a Python interpreter with pytest installed. "
                    "Try running: python -m pip install -r requirements.txt\n"
                )
                log_file.write(missing_message)
                log_file.flush()
                _broadcast_log_line(missing_message)
                _save_run_state(
                    {
                        "suite_id": suite_id,
                        "suite_label": suite["label"],
                        "suite_method": suite["method"],
                        "command": None,
                        "started_at": started_at,
                        "completed_at": datetime.now().astimezone().isoformat(timespec="seconds"),
                        "status": "runner_missing",
                        "exit_code": None,
                    "message": missing_message.strip(),
                    "linked_xml_filename": None,
                    "linked_html_filename": None,
                    "linked_excel_filename": None,
                    "report_found": False,
                    "paused": False,
                    "pid": None,
                }
            )
                _broadcast_log_line("__RUN_COMPLETE__")
                return

            command = [*command_prefix, *suite["args"]]
            process_env = os.environ.copy()
            process_env["DASHBOARD_HEADLESS"] = "1"
            process_env["PYTEST_RUNNING_FROM_DASHBOARD"] = "1"
            start_message = (
                f"[dashboard] Starting suite '{suite['label']}' "
                f"({suite['method']}) in headless background mode with: {' '.join(command)}\n"
            )
            log_file.write(start_message)
            log_file.flush()
            _broadcast_log_line(start_message)
            _save_run_state(
                {
                    "suite_id": suite_id,
                    "suite_label": suite["label"],
                    "suite_method": suite["method"],
                    "suite_description": suite["description"],
                    "browser_mode": "Headless background (no popup browser)",
                    "command": " ".join(command),
                    "command_preview": _build_suite_command_preview(suite),
                    "started_at": started_at,
                    "completed_at": None,
                    "status": "running",
                    "exit_code": None,
                    "message": "Pytest suite is running.",
                    "linked_xml_filename": None,
                    "linked_html_filename": None,
                    "linked_excel_filename": None,
                    "report_found": False,
                    "paused": False,
                    "pid": None,
                }
            )
            process = subprocess.Popen(
                command,
                cwd=PROJECT_ROOT,
                env=process_env,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                bufsize=1,
            )
            _active_process = process
            _merge_run_state(pid=process.pid)

            assert process.stdout is not None
            for line in process.stdout:
                log_file.write(line)
                log_file.flush()
                _broadcast_log_line(line)

            return_code = process.wait()
            linked_run = _pick_generated_run(before_xml_files)
            linked_xml_filename = linked_run["xml_filename"] if linked_run else None
            linked_html_filename = linked_run["html_filename"] if linked_run else None
            linked_excel_filename = linked_run["excel_filename"] if linked_run else None

            if return_code == 0 and linked_run:
                state_status = "passed"
                state_message = "Suite finished successfully and produced a fresh report."
            elif return_code == 1 and linked_run:
                state_status = "test_failures"
                state_message = "Suite finished with real failing test cases."
            elif return_code == 5:
                state_status = "no_tests"
                state_message = "Pytest completed without collecting any tests for the selected suite."
            elif linked_run:
                state_status = "infra_error"
                state_message = "Pytest ended with a non-standard exit code. A report was generated but the run needs attention."
            else:
                state_status = "infra_error"
                state_message = "Pytest did not produce a fresh XML report for this run. Dashboard will not treat old reports as new results."

            done_message = f"[dashboard] Pytest finished with exit code {return_code}.\n"
            log_file.write(done_message)
            log_file.flush()
            _broadcast_log_line(done_message)
            if linked_run:
                linked_message = f"[dashboard] Linked report: {linked_xml_filename}\n"
                log_file.write(linked_message)
                log_file.flush()
                _broadcast_log_line(linked_message)

            _save_run_state(
                {
                    "suite_id": suite_id,
                    "suite_label": suite["label"],
                    "suite_method": suite["method"],
                    "suite_description": suite["description"],
                    "browser_mode": "Headless background (no popup browser)",
                    "command": " ".join(command),
                    "command_preview": _build_suite_command_preview(suite),
                    "started_at": started_at,
                    "completed_at": datetime.now().astimezone().isoformat(timespec="seconds"),
                    "status": state_status,
                    "exit_code": return_code,
                    "message": state_message,
                    "linked_xml_filename": linked_xml_filename,
                    "linked_html_filename": linked_html_filename,
                    "linked_excel_filename": linked_excel_filename,
                    "report_found": bool(linked_run),
                    "tests": linked_run["tests"] if linked_run else None,
                    "passed": linked_run["passed"] if linked_run else None,
                    "failed": (linked_run["failures"] + linked_run["errors"]) if linked_run else None,
                    "paused": False,
                    "pid": None,
                }
            )
            _broadcast_log_line("__RUN_COMPLETE__")
    finally:
        with _run_lock:
            _active_process = None
            _run_in_progress = False
            _run_paused = False


@app.route("/")
def index() -> Response:
    return redirect(url_for("dashboard"))


@app.route("/dashboard")
def dashboard() -> str:
    runs = get_all_runs()
    latest_run = runs[0] if runs else None
    history_rows = runs[:10]
    historical = get_historical_data(runs, n=7)
    module_trend = get_module_trend_data(runs, n=10)
    last_run_state = _refresh_live_run_state()
    live_pid = _get_live_run_pid()
    totals = {
        "runs": len(runs),
        "tests": sum(run["tests"] for run in runs),
        "failures": sum(run["failures"] + run["errors"] for run in runs),
        "duration": format_duration(sum(run["time_seconds"] for run in runs)),
    }
    return render_template(
        "dashboard.html",
        latest_run=latest_run,
        history_rows=history_rows,
        historical=historical,
        module_trend=module_trend,
        totals=totals,
        suite_running=live_pid is not None,
        suite_paused=live_pid is not None and bool((last_run_state or {}).get("paused")),
        test_suites=_get_suite_choices(),
        selected_suite_id=(last_run_state or {}).get("suite_id", "full"),
        last_run_state=last_run_state,
    )


@app.route("/project-doc")
def project_doc() -> str:
    return render_template("project_doc.html", **_get_project_doc_context())


@app.route("/excel-reports")
def excel_reports() -> str:
    excel_files = _get_excel_reports()
    latest_excel = excel_files[0] if excel_files else None
    files_with_failures = sum(1 for item in excel_files if item["failed_cases"] > 0)
    total_failed_cases = sum(item["failed_cases"] for item in excel_files)
    return render_template(
        "excel_reports.html",
        excel_files=excel_files,
        latest_excel=latest_excel,
        total_excel_files=len(excel_files),
        files_with_failures=files_with_failures,
        total_failed_cases=total_failed_cases,
    )


@app.route("/reports")
def reports_latest() -> str:
    runs = get_all_runs()
    if not runs:
        abort(404)
    last_run_state = _load_run_state()
    selected_run = None
    if last_run_state and last_run_state.get("linked_xml_filename"):
        selected_run = _get_run_by_filename(last_run_state["linked_xml_filename"])
    if selected_run is None:
        selected_run = runs[0]
    selected_case_index = _resolve_selected_case_index(selected_run, request.args.get("case"))
    selected_case_sheet = build_case_sheet(selected_run, selected_run["testcases"][selected_case_index])
    return render_template(
        "reports.html",
        selected_case_index=selected_case_index,
        selected_case_sheet=selected_case_sheet,
        **_build_reports_context(selected_run, runs, last_run_state=last_run_state),
    )


@app.route("/reports/<xml_filename>")
def reports_by_file(xml_filename: str) -> str:
    runs = get_all_runs()
    selected_run = _find_run(xml_filename)
    if selected_run is None:
        abort(404)
    selected_case_index = _resolve_selected_case_index(selected_run, request.args.get("case"))
    selected_case_sheet = build_case_sheet(selected_run, selected_run["testcases"][selected_case_index])
    return render_template(
        "reports.html",
        selected_case_index=selected_case_index,
        selected_case_sheet=selected_case_sheet,
        **_build_reports_context(selected_run, runs, last_run_state=_load_run_state()),
    )


@app.route("/logs")
def logs() -> str:
    if LAST_LOG_FILE.exists():
        content = LAST_LOG_FILE.read_text(encoding="utf-8", errors="replace")
    else:
        content = "No dashboard-triggered run has been executed yet."
    last_run_state = _refresh_live_run_state()
    live_pid = _get_live_run_pid()
    return render_template(
        "logs.html",
        log_content=content,
        suite_running=live_pid is not None,
        suite_paused=live_pid is not None and bool((last_run_state or {}).get("paused")),
        test_suites=_get_suite_choices(),
        selected_suite_id=(last_run_state or {}).get("suite_id", "full"),
        last_run_state=last_run_state,
    )


@app.route("/download/<html_filename>")
def download_report(html_filename: str):
    report_path = _safe_report_path(html_filename)
    return send_from_directory(REPORTS_DIR, report_path.name, as_attachment=True)


@app.route("/download-excel/<xlsx_filename>")
def download_excel_report(xlsx_filename: str):
    excel_path = _safe_excel_path(xlsx_filename)
    return send_from_directory(REPORTS_DIR, excel_path.name, as_attachment=True)


@app.route("/report-artifacts/<path:artifact_path>")
def report_artifact(artifact_path: str):
    artifact_file = _safe_report_artifact_path(artifact_path)
    relative_path = str(artifact_file.relative_to(REPORTS_DIR)).replace("\\", "/")
    return send_from_directory(REPORTS_DIR, relative_path)


@app.route("/download-cumulative-excel")
def download_cumulative_excel():
    output_path = generate_cumulative_report(reports_dir=REPORTS_DIR)
    return send_from_directory(
        REPORTS_DIR,
        output_path.name,
        as_attachment=True,
        download_name=output_path.name,
    )


@app.route("/api/cumulative-summary")
def api_cumulative_summary():
    from catalog.test_case_meta import TEST_CASE_META as meta_dict

    latest = build_latest_results(REPORTS_DIR)
    return get_coverage_summary(latest, meta_dict)


@app.route("/download-pdf")
@app.route("/download-pdf/<xml_filename>")
def download_pdf_report(xml_filename: str | None = None):
    runs = get_all_runs()
    if not runs:
        abort(404)
    selected_run = _find_run(xml_filename) if xml_filename else runs[0]
    if selected_run is None:
        abort(404)

    timestamp_token = selected_run["xml_filename"].removeprefix("results_").removesuffix(".xml")
    output_path = REPORTS_DIR / f"report_pdf_{timestamp_token}.pdf"
    critical_failures = get_critical_failures(selected_run)
    automated_insights = get_automated_insights(selected_run, runs)
    generate_pdf_report(selected_run, critical_failures, automated_insights, output_path)
    return send_from_directory(REPORTS_DIR, output_path.name, as_attachment=True, download_name=output_path.name)


@app.route("/download-word-report")
def download_word_report():
    from catalog.test_case_meta import TEST_CASE_META as meta_dict

    runs = get_all_runs()
    latest_run = runs[0] if runs else None
    latest = build_latest_results(REPORTS_DIR)
    cumulative_coverage = get_coverage_summary(latest, meta_dict)
    bug_rows = get_bug_tracker_rows(runs)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = REPORTS_DIR / f"bao_cao_word_tu_dong_{timestamp}.docx"
    generate_word_report(
        output_path=output_path,
        latest_run=latest_run,
        cumulative_coverage=cumulative_coverage,
        bugs=bug_rows,
        project_doc=PROJECT_WORD_DATA,
    )
    return send_from_directory(REPORTS_DIR, output_path.name, as_attachment=True, download_name=output_path.name)


@app.route("/bugs")
def bugs() -> str:
    runs = get_all_runs()
    bug_rows = get_bug_tracker_rows(runs)
    severity_totals = {
        "HIGH": sum(1 for row in bug_rows if row["severity"] == "HIGH"),
        "MEDIUM": sum(1 for row in bug_rows if row["severity"] == "MEDIUM"),
        "LOW": sum(1 for row in bug_rows if row["severity"] == "LOW"),
    }
    return render_template(
        "bugs.html",
        bug_rows=bug_rows,
        severity_totals=severity_totals,
        total_bugs=len(bug_rows),
    )


@app.route("/compare")
@app.route("/compare/<xml_a>/<xml_b>")
def compare(xml_a: str | None = None, xml_b: str | None = None) -> str:
    runs = get_all_runs()
    if len(runs) < 2:
        abort(404)

    default_a = runs[1]["xml_filename"] if len(runs) > 1 else runs[0]["xml_filename"]
    default_b = runs[0]["xml_filename"]
    xml_a = xml_a or request.args.get("xml_a") or default_a
    xml_b = xml_b or request.args.get("xml_b") or default_b

    run_a = _find_run(xml_a)
    run_b = _find_run(xml_b)
    if run_a is None or run_b is None:
        abort(404)

    comparison = compare_runs(run_a, run_b)
    return render_template(
        "compare.html",
        all_runs=runs,
        comparison=comparison,
        selected_a=xml_a,
        selected_b=xml_b,
    )


@app.route("/run-suite", methods=["POST"])
def run_suite():
    global _run_in_progress, _run_paused
    payload = request.get_json(silent=True) or request.form.to_dict()
    suite_id, suite = _get_suite(payload.get("suite_id"))
    with _run_lock:
        if _run_in_progress:
            return {"status": "busy", "message": "A pytest run is already in progress."}, 409

        while not _log_queue.empty():
            try:
                _log_queue.get_nowait()
            except queue.Empty:
                break

        _run_paused = False
        _save_run_state(
            {
                "suite_id": suite_id,
                "suite_label": suite["label"],
                "suite_method": suite["method"],
                "suite_description": suite["description"],
                "browser_mode": "Headless background (no popup browser)",
                "command": None,
                "command_preview": _build_suite_command_preview(suite),
                "started_at": datetime.now().astimezone().isoformat(timespec="seconds"),
                "completed_at": None,
                "status": "running",
                "exit_code": None,
                "message": f"Suite '{suite['label']}' is queued and starting.",
                "linked_xml_filename": None,
                "linked_html_filename": None,
                "linked_excel_filename": None,
                "report_found": False,
                "paused": False,
                "pid": None,
            }
        )
        _run_in_progress = True
        thread = threading.Thread(target=_run_pytest_worker, args=(suite_id,), daemon=True)
        thread.start()

    return {
        "status": "started",
        "message": f"Pytest suite '{suite['label']}' started.",
        "suite_id": suite_id,
        "suite_label": suite["label"],
        "command_preview": _build_suite_command_preview(suite),
    }, 202


@app.route("/toggle-suite-pause", methods=["POST"])
def toggle_suite_pause():
    last_run_state = _refresh_live_run_state()
    if psutil is None:
        return {
            "status": "unsupported",
            "message": "Pause/resume requires psutil. Run: python -m pip install -r requirements.txt",
        }, 501

    with _run_lock:
        if _get_live_run_pid() is None:
            return {
                "status": "idle",
                "message": last_run_state.get("message", "No active pytest run to pause or resume.")
                if last_run_state else "No active pytest run to pause or resume.",
            }, 409

        effective_paused = _run_paused or bool((last_run_state or {}).get("paused"))
        if effective_paused:
            resumed = _resume_active_process_tree()
            if not resumed:
                return {"status": "error", "message": "Could not resume the active pytest process."}, 500
            _broadcast_log_line("[dashboard] Suite resumed from browser control.\n")
            return {"status": "running", "paused": False, "message": "Suite resumed."}, 200

        paused = _suspend_active_process_tree()
        if not paused:
            return {"status": "error", "message": "Could not pause the active pytest process."}, 500
        _broadcast_log_line("[dashboard] Suite paused from browser control.\n")
        return {"status": "paused", "paused": True, "message": "Suite paused."}, 200


@app.route("/stream-log")
def stream_log() -> Response:
    def event_stream():
        while True:
            try:
                line = _log_queue.get(timeout=30)
            except queue.Empty:
                yield "event: heartbeat\ndata: keep-alive\n\n"
                continue

            payload = json.dumps({"line": line})
            if line == "__RUN_COMPLETE__":
                yield f"event: done\ndata: {payload}\n\n"
                break

            yield f"event: log\ndata: {payload}\n\n"

    return Response(event_stream(), mimetype="text/event-stream")


if __name__ == "__main__":
    _open_browser_on_start()
    app.run(host="127.0.0.1", port=5050, debug=False)
